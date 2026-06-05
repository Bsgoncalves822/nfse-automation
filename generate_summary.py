import os
import glob
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_summary(filter_names=None, filter_month=None):
    settings_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config', 'settings.json')
    with open(settings_path, encoding='utf-8') as f:
        settings = json.load(f)

    base         = Path(settings['downloads_path'])
    empresas_dir = base / 'Empresas'
    out_path     = empresas_dir / 'resumo_nfse.xlsx'
    out_path.parent.mkdir(parents=True, exist_ok=True)

    rows = []

    for company_dir in sorted(empresas_dir.iterdir()):
        if not company_dir.is_dir():
            continue
        if filter_names and not any(company_dir.name.startswith(n) or n in company_dir.name for n in filter_names):
            continue
        for month_dir in sorted(company_dir.iterdir()):
            if not month_dir.is_dir():
                continue
            if filter_month and month_dir.name != filter_month:
                continue
            federal_xml_dir = month_dir / 'federal' / 'xmls'
            count = len(glob.glob(str(federal_xml_dir / '*.xml'))) if federal_xml_dir.exists() else 0
            rows.append({
                'empresa': company_dir.name,
                'periodo': month_dir.name,
                'count':   count,
            })

    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumo NFS-e'

    title_font  = Font(name='Arial', bold=True, size=11, color='1A56A0')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', start_color='1A56A0')
    total_font  = Font(name='Arial', bold=True, size=10)
    total_fill  = PatternFill('solid', start_color='D6E4F7')
    normal_font = Font(name='Arial', size=10)
    center      = Alignment(horizontal='center', vertical='center')
    left        = Alignment(horizontal='left',   vertical='center')
    thin        = Side(style='thin', color='BFBFBF')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:C1')
    ws['A1']           = f'Resumo NFS-e com Retencao Federal — gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A1'].font      = title_font
    ws['A1'].alignment = left
    ws.row_dimensions[1].height = 20

    headers = ['Nome da Empresa', 'Periodo (MM-YYYY)', 'Notas com Retencao Federal (XMLs)']
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[2].height = 16

    for i, r in enumerate(rows, start=3):
        fill = PatternFill('solid', start_color='F0F5FC') if i % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
        for col, val in enumerate([r['empresa'], r['periodo'], r['count']], 1):
            cell           = ws.cell(row=i, column=col, value=val)
            cell.font      = normal_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = center if col > 1 else left

    total_row = len(rows) + 3
    ws.cell(row=total_row, column=1, value='TOTAL').font      = total_font
    ws.cell(row=total_row, column=1).fill                     = total_fill
    ws.cell(row=total_row, column=1).alignment                = left
    ws.cell(row=total_row, column=1).border                   = border
    total_cell           = ws.cell(row=total_row, column=3, value=sum(r['count'] for r in rows))
    total_cell.font      = total_font
    total_cell.fill      = total_fill
    total_cell.alignment = center
    total_cell.border    = border
    for col in [2]:
        cell        = ws.cell(row=total_row, column=col)
        cell.fill   = total_fill
        cell.border = border

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.freeze_panes = 'A3'

    wb.save(str(out_path))
    print(f'[OK] Resumo salvo: {out_path}', flush=True)
    return str(out_path)

if __name__ == '__main__':
    generate_summary()
