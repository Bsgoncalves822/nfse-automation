import os
import json
import sys
from pathlib import Path
from datetime import datetime
from playwright.sync_api import sync_playwright

APP_URL      = "https://script.google.com/macros/s/AKfycbxIVVGkPUdOeNaMa-IJosK3wFvKQbpzDd6l4Sze9scdahdf2i6fD0hWMYubseuDjAnV/exec"
USERNAME     = "mario"
PASSWORD     = "123"
GOOGLE_EMAIL = "orprocondigital@gmail.com"
GOOGLE_PASS  = "123Orprocon2025"
AUTH_FILE    = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config', 'integracao_auth.json')

def read_resumo(resumo_path, month):
    try:
        from openpyxl import load_workbook
        wb     = load_workbook(resumo_path, read_only=True, data_only=True)
        ws     = wb.active
        result = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            name   = str(row[0]).strip() if row[0] else ''
            period = str(row[1]).strip() if row[1] else ''
            count  = int(row[2]) if row[2] and str(row[2]).isdigit() else 0
            if not name or name == 'TOTAL' or name == 'None':
                continue
            if period != month:
                continue
            result[name] = count
        print(f'[OK] Resumo lido: {len(result)} empresa(s) para {month}')
        return result
    except Exception as e:
        print(f'[ERRO] Falha ao ler resumo: {e}')
        return {}

def load_companies_map():
    try:
        import urllib.request, csv, io
        url = 'https://docs.google.com/spreadsheets/d/1MI4xI6rSWfYVYTtPfXOzNPon-AGq0KXh/export?format=csv'
        with urllib.request.urlopen(url) as r:
            content = r.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        return {row['name'].strip(): row['cnpj'].strip() for row in reader if row.get('cnpj')}
    except Exception as e:
        print(f'[AVISO] Google Sheets falhou: {e}, usando companies.json')
        p = Path(__file__).parent / 'config' / 'companies.json'
        with open(p, encoding='utf-8') as f:
            companies = json.load(f)
        return {c['name']: c['cnpj'] for c in companies}

def do_google_login(page):
    """Perform Google login and return True if successful."""
    try:
        page.wait_for_timeout(2000)
        if 'accounts.google.com' not in page.url:
            return True

        print('[OK] Fazendo login no Google...')
        page.wait_for_selector('input[type="email"]', timeout=15000)
        page.fill('input[type="email"]', GOOGLE_EMAIL)
        page.wait_for_timeout(500)

        next_btn = page.query_selector('#identifierNext') or page.query_selector('button:has-text("Next")')
        if next_btn:
            next_btn.click()
        else:
            page.keyboard.press('Enter')

        page.wait_for_timeout(3000)
        page.wait_for_selector('input[type="password"]', timeout=15000)
        page.fill('input[type="password"]', GOOGLE_PASS)
        page.wait_for_timeout(500)

        pass_btn = page.query_selector('#passwordNext') or page.query_selector('button:has-text("Next")')
        if pass_btn:
            pass_btn.click()
        else:
            page.keyboard.press('Enter')

        page.wait_for_load_state('networkidle', timeout=30000)
        page.wait_for_timeout(3000)
        print('[OK] Login Google concluido')
        return True
    except Exception as e:
        print(f'[ERRO] Login Google falhou: {e}')
        return False

def do_reinf_login(page):
    """Perform EFD-REINF login if form is visible."""
    try:
        if page.is_visible('input[placeholder="Seu usuario"]'):
            page.fill('input[placeholder="Seu usuario"]', USERNAME)
            page.fill('input[placeholder="Sua senha"]', PASSWORD)
            page.click('#blg')
            page.wait_for_load_state('networkidle', timeout=30000)
            page.wait_for_timeout(2000)
            print('[OK] Login EFD-REINF efetuado')
    except Exception as e:
        print(f'[AVISO] Login EFD-REINF: {e}')

def run_integracao(resumo_path, month):
    resumo   = read_resumo(resumo_path, month)
    name_map = load_companies_map()

    if not resumo:
        print(f'[INFO] Nenhuma empresa no resumo para {month}')
        return {'updated': 0, 'sem_movimento': 0, 'not_found': 0, 'error': 0}

    cnpj_map = {}
    for name, count in resumo.items():
        cnpj = name_map.get(name)
        if cnpj:
            cnpj_map[cnpj] = count
        else:
            print(f'[AVISO] CNPJ nao encontrado para: {name}')

    print(f'[OK] {len(cnpj_map)} empresa(s) mapeadas')

    try:
        dt = datetime.strptime(month, '%m-%Y')
    except:
        print(f'[ERRO] Formato de mes invalido: {month}')
        return {}

    pt_months = {
        1:'Janeiro', 2:'Fevereiro', 3:'Marco', 4:'Abril',
        5:'Maio', 6:'Junho', 7:'Julho', 8:'Agosto',
        9:'Setembro', 10:'Outubro', 11:'Novembro', 12:'Dezembro'
    }
    month_pt = pt_months[dt.month]
    year_str = str(dt.year)

    stats = {'updated': 0, 'sem_movimento': 0, 'not_found': 0, 'error': 0}

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, args=['--start-maximized'])

        # Load saved auth state if exists
        if os.path.exists(AUTH_FILE):
            context = browser.new_context(storage_state=AUTH_FILE)
            print('[OK] Sessao salva carregada')
        else:
            context = browser.new_context()
            print('[OK] Nova sessao iniciada')

        page = context.new_page()

        try:
            print('[OK] Abrindo EFD-REINF...')
            page.goto(APP_URL, wait_until='networkidle', timeout=60000)

            # Handle Google login if needed
            if 'accounts.google.com' in page.url:
                if not do_google_login(page):
                    print('[ERRO] Nao foi possivel fazer login no Google')
                    return stats
                page.goto(APP_URL, wait_until='networkidle', timeout=60000)
                page.wait_for_timeout(2000)

            # Handle EFD-REINF login if needed
            do_reinf_login(page)

            # Save auth state for future runs
            context.storage_state(path=AUTH_FILE)
            print('[OK] Sessao salva para proxima execucao')

            # Navigate to Lancamentos
            page.click('button[data-v="lancamentos"]')
            page.wait_for_timeout(2000)
            print('[OK] Lancamentos carregado')

            # Select month and year
            selects = page.query_selector_all('select')
            if len(selects) >= 2:
                selects[0].select_option(label=month_pt)
                selects[1].select_option(value=year_str)
                page.wait_for_timeout(2000)
                print(f'[OK] Periodo selecionado: {month_pt} {year_str}')

            # Process each company
            for cnpj, count in cnpj_map.items():
                try:
                    search = page.query_selector('input[placeholder="Buscar empresa ou CNPJ..."]')
                    if search:
                        search.triple_click()
                        search.type(cnpj, delay=50)
                        page.wait_for_timeout(1000)

                    if count > 0:
                        cell = page.query_selector(f'td[data-c="{cnpj}"][data-f="qtdNF"]')
                        if not cell:
                            print(f'[AVISO] Nao encontrada: {cnpj}')
                            stats['not_found'] += 1
                            continue

                        cell.click()
                        page.wait_for_timeout(400)
                        page.keyboard.press('Control+a')
                        page.keyboard.type(str(count))
                        page.keyboard.press('Enter')
                        page.wait_for_timeout(600)

                        fiscal_btn = page.query_selector(f'button[data-c="{cnpj}"][data-n="fiscal"]')
                        if fiscal_btn:
                            fiscal_btn.click()
                            page.wait_for_timeout(600)
                            print(f'[OK] {cnpj} — {count} nota(s) -> Fiscal')
                            stats['updated'] += 1
                        else:
                            print(f'[AVISO] Botao Fiscal nao encontrado: {cnpj}')
                            stats['error'] += 1

                    else:
                        checkbox = page.query_selector(f'input[data-c="{cnpj}"][data-f="semMovimento"]')
                        if checkbox:
                            if not checkbox.is_checked():
                                checkbox.click()
                                page.wait_for_timeout(400)
                            print(f'[OK] {cnpj} — Sem movimento')
                            stats['sem_movimento'] += 1
                        else:
                            print(f'[AVISO] Checkbox S/M nao encontrado: {cnpj}')
                            stats['not_found'] += 1

                except Exception as e:
                    print(f'[ERRO] {cnpj}: {e}')
                    stats['error'] += 1

                finally:
                    search = page.query_selector('input[placeholder="Buscar empresa ou CNPJ..."]')
                    if search:
                        search.triple_click()
                        search.press('Delete')
                        page.wait_for_timeout(200)

        except Exception as e:
            print(f'[ERRO] Falha geral: {e}')
        finally:
            page.close()
            context.close()
            browser.close()

    print(f'\n{"="*50}')
    print(f'Integracao concluida: {month}')
    print(f'  Atualizadas:     {stats["updated"]}')
    print(f'  Sem movimento:   {stats["sem_movimento"]}')
    print(f'  Nao encontradas: {stats["not_found"]}')
    print(f'  Erros:           {stats["error"]}')
    return stats

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Uso: python integracao.py <resumo_path> <MM-YYYY>')
        sys.exit(1)
    run_integracao(sys.argv[1], sys.argv[2])