import os
import re
import time
import shutil
from datetime import datetime
from collections import defaultdict
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_URL = "https://www.nfse.gov.br/EmissorNacional"

# ─────────────────────────────────────────────
# Namespace-agnostic XML helpers
# ─────────────────────────────────────────────

def _local(tag):
    return tag.split('}')[-1] if '}' in tag else tag

def find_text(root, tag):
    for el in root.iter():
        if _local(el.tag) == tag and el.text:
            return el.text.strip()
    return ''

def find_float(root, tag):
    val = find_text(root, tag)
    try:
        return float(val.replace(',', '.'))
    except:
        return 0.0

def find_all_text(root, tag):
    return [el.text.strip() for el in root.iter()
            if _local(el.tag) == tag and el.text]

# ─────────────────────────────────────────────
# XML classification
# ─────────────────────────────────────────────

FEDERAL_FIELDS = ['vRetIRRF','vRetCSLL','vPis','vCofins','vRetINSS','vRetCP']

def is_federal(xml_path):
    try:
        root = ET.parse(xml_path).getroot()
        for field in FEDERAL_FIELDS:
            try:
                if find_float(root, field) > 0:
                    return True
            except:
                pass
        return False
    except:
        return False

def is_municipal(xml_path):
    try:
        root = ET.parse(xml_path).getroot()
        return find_text(root, 'tpRetISSQN') == '1'
    except:
        return False

def get_nnfse(xml_path):
    try:
        return find_text(ET.parse(xml_path).getroot(), 'nNFSe') or None
    except:
        return None

def parse_xml_full(xml_path):
    try:
        root = ET.parse(xml_path).getroot()

        n_nfse      = find_text(root, 'nNFSe')
        x_loc_emi   = find_text(root, 'xLocEmi')
        x_loc_prest = find_text(root, 'xLocPrestacao') or x_loc_emi
        dh_emi      = find_text(root, 'dhEmi')
        x_trib_nac  = find_text(root, 'xTribNac')
        c_trib_nac  = find_text(root, 'cTribNac')
        x_desc_serv = find_text(root, 'xDescServ')
        c_stat      = find_text(root, 'cStat')

        # Emitente — try emit first, then prest
        cnpj_emit = find_text(root, 'CNPJ')  # first CNPJ in doc is emitente
        nome_emit = find_text(root, 'xNome')

        # Tomador
        # Find toma element and get its CNPJ/xNome
        cnpj_toma = ''
        nome_toma = ''
        for el in root.iter():
            if _local(el.tag) == 'toma':
                cnpj_toma = find_text(el, 'CNPJ')
                nome_toma = find_text(el, 'xNome')
                break

        def fmt_cnpj(c):
            c = ''.join(filter(str.isdigit, c))
            if len(c) == 14:
                return f'{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}'
            elif len(c) == 11:
                return f'{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}'
            return c

        try:
            dt = datetime.fromisoformat(dh_emi[:19])
            data_fmt = dt.strftime('%d/%m/%Y')
        except:
            data_fmt = dh_emi[:10]

        v_serv     = find_float(root, 'vServ') or find_float(root, 'vLiq')
        v_ret_irrf = find_float(root, 'vRetIRRF')
        v_ret_csll = find_float(root, 'vRetCSLL')
        tp_pis     = find_text(root, 'tpRetPisCofins')
        v_pis      = find_float(root, 'vPis') if tp_pis == '1' else 0.0
        v_cofins   = find_float(root, 'vCofins') if tp_pis == '1' else 0.0
        v_ret_inss = find_float(root, 'vRetINSS')
        v_ret_cp   = find_float(root, 'vRetCP')
        v_issqn    = find_float(root, 'vISSQN')
        v_cbs      = find_float(root, 'vCBS')
        v_ibs      = find_float(root, 'vIBSTot')

        tp_ret_iss = find_text(root, 'tpRetISSQN')
        iss_ret    = v_issqn if tp_ret_iss == '1' else 0.0
        inss_val   = v_ret_cp if v_ret_cp > 0 else v_ret_inss
        total_ret  = v_ret_irrf + v_ret_csll + v_pis + v_cofins + inss_val + v_cbs + v_ibs
        v_liq      = find_float(root, 'vLiq') or v_serv
        cancelada  = c_stat not in ('100', '')
        is_fed     = total_ret > 0
        is_mun     = (not is_fed) and (tp_ret_iss == '1')

        return {
            'numero':       n_nfse,
            'local_prest':  x_loc_prest,
            'emissao':      data_fmt,
            'cnpj_emit':    fmt_cnpj(cnpj_emit),
            'nome_emit':    nome_emit,
            'cnpj_toma':    fmt_cnpj(cnpj_toma),
            'nome_toma':    nome_toma,
            'nbs':          find_text(root, 'cNBS'),
            'desc_nbs':     find_text(root, 'xNBS'),
            'cod_trib':     c_trib_nac,
            'desc_trib':    x_trib_nac[:90] if x_trib_nac else '',
            'desc_serv':    x_desc_serv[:200] if x_desc_serv else '',
            'v_serv':       v_serv,
            'iss_aliq':     '',
            'iss_val':      v_issqn,
            'iss_ret':      iss_ret,
            'pis_ret':      v_pis,
            'cofins_ret':   v_cofins,
            'ir_ret':       v_ret_irrf,
            'csll_ret':     v_ret_csll,
            'inss_ret':     inss_val,
            'cbs_ret':      v_cbs,
            'ibs_ret':      v_ibs,
            'total_ret':    total_ret,
            'v_liq':        v_liq,
            'is_federal':   is_fed,
            'is_municipal': is_mun,
            'cancelada':    cancelada,
            'obs':          'CANCELADA - NFS-e cancelada' if cancelada else '',
        }
    except Exception as e:
        print(f'[AVISO] Erro ao parsear {xml_path}: {e}', flush=True)
        return None

# ─────────────────────────────────────────────
# Excel report
# ─────────────────────────────────────────────

def generate_recebidas_excel(rows, company_name, month, download_dir):
    wb = Workbook()
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill('solid', start_color='1A56A0')
    tot_font = Font(name='Arial', bold=True, size=10)
    tot_fill = PatternFill('solid', start_color='D6E4F7')
    nrm_font = Font(name='Arial', size=10)
    ttl_font = Font(name='Arial', bold=True, size=11, color='1A56A0')
    ctr  = Alignment(horizontal='center', vertical='center')
    lft  = Alignment(horizontal='left',   vertical='center')
    rgt  = Alignment(horizontal='right',  vertical='center')
    thin = Side(style='thin', color='BFBFBF')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    money = '#,##0.00'

    ret_fed_rows = [r for r in rows if r['total_ret'] > 0]
    ret_mun_rows = [r for r in rows if r['total_ret'] == 0 and r['iss_ret'] > 0]
    ret_rows     = [r for r in rows if r['total_ret'] > 0 or r['iss_ret'] > 0]
    cancel_rows = [r for r in rows if r['cancelada']]

    cols1 = ['Nr NFSe','Local Prest','Emissao','CNPJ/CPF Emitente','Razao Emitente',
             'CNPJ/CPF Tomador','Razao Tomador','NBS','Desc. NBS','Cod. Tributacao',
             'Descr. Tributacao','Descr. Servico','Vl. Servico','ISS','ISS Valor',
             'ISS Ret.','Pis Ret.','Cofins Ret.','IR Ret.','CSLL Ret.','INSS Ret.',
             'CBS Ret.','IBS Ret.','Vl. Liquido','Observacao']
    money_cols1 = {13,15,16,17,18,19,20,21,22,23,24}

    ws1 = wb.active
    ws1.title = 'NFS-e'
    ws1.merge_cells('A1:Y1')
    ws1['A1'] = company_name
    ws1['A1'].font = ttl_font; ws1['A1'].alignment = lft
    ws1.merge_cells('A2:Y2')
    ws1['A2'] = f"Relatorio de Notas Recebidas   Gerado em: {datetime.now().strftime('%d/%m/%Y as %H:%M:%S')}"
    ws1['A2'].font = Font(name='Arial', size=9, color='7F8C8D')
    for c, h in enumerate(cols1, 1):
        cell = ws1.cell(row=3, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = ctr; cell.border = bdr
    for i, r in enumerate(rows, start=4):
        fill = PatternFill('solid', start_color='F0F5FC') if i%2==0 else PatternFill('solid', start_color='FFFFFF')
        vals = [r['numero'],r['local_prest'],r['emissao'],r['cnpj_emit'],r['nome_emit'],
                r['cnpj_toma'],r['nome_toma'],r['nbs'],r['desc_nbs'],r['cod_trib'],
                r['desc_trib'],r['desc_serv'],r['v_serv'],r['iss_aliq'],r['iss_val'],
                r['iss_ret'],r['pis_ret'],r['cofins_ret'],r['ir_ret'],r['csll_ret'],
                r['inss_ret'],r['cbs_ret'],r['ibs_ret'],r['v_liq'],r['obs']]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=c, value=v)
            cell.font = nrm_font; cell.fill = fill; cell.border = bdr
            if c in money_cols1: cell.number_format = money; cell.alignment = rgt
            elif c in {1,2,3}: cell.alignment = ctr
            else: cell.alignment = lft
    tr = len(rows)+4
    ws1.cell(row=tr, column=1, value='TOTAIS').font = tot_font
    ws1.cell(row=tr, column=1).fill = tot_fill
    ws1.cell(row=tr, column=1).alignment = lft
    ws1.cell(row=tr, column=1).border = bdr
    tot_map1 = {13:'v_serv',15:'iss_val',16:'iss_ret',17:'pis_ret',18:'cofins_ret',
                19:'ir_ret',20:'csll_ret',21:'inss_ret',22:'cbs_ret',23:'ibs_ret',24:'v_liq'}
    for c in range(2, len(cols1)+1):
        cell = ws1.cell(row=tr, column=c); cell.fill = tot_fill; cell.border = bdr
        if c in tot_map1:
            cell.value = sum(r[tot_map1[c]] for r in rows)
            cell.number_format = money; cell.alignment = rgt; cell.font = tot_font
    ws1.column_dimensions['A'].width = 15; ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 40; ws1.column_dimensions['K'].width = 50
    ws1.column_dimensions['L'].width = 50; ws1.freeze_panes = 'A4'

    cols2 = ['Nr NFSe','Emissao','CNPJ/CPF Emitente','Razao Emitente',
             'Vl. Servico','ISS Ret.','Pis Ret.','Cofins Ret.',
             'IR Ret.','CSLL Ret.','INSS Ret.','CBS Ret.','IBS Ret.','Total Retido']
    money_cols2 = {5,6,7,8,9,10,11,12,13,14}
    ws2 = wb.create_sheet('Retencao Federal')
    for c, h in enumerate(cols2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = ctr; cell.border = bdr
    for i, r in enumerate(ret_fed_rows, start=2):
        fill = PatternFill('solid', start_color='F0F5FC') if i%2==0 else PatternFill('solid', start_color='FFFFFF')
        tot = r['total_ret'] + r['iss_ret']
        vals = [r['numero'],r['emissao'],r['cnpj_emit'],r['nome_emit'],
                r['v_serv'],r['iss_ret'],r['pis_ret'],r['cofins_ret'],
                r['ir_ret'],r['csll_ret'],r['inss_ret'],r['cbs_ret'],r['ibs_ret'],tot]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.font = nrm_font; cell.fill = fill; cell.border = bdr
            if c in money_cols2: cell.number_format = money; cell.alignment = rgt
            elif c in {1,2}: cell.alignment = ctr
            else: cell.alignment = lft
    tr2 = len(ret_fed_rows)+2
    ws2.cell(row=tr2, column=1, value='TOTAIS').font = tot_font
    ws2.cell(row=tr2, column=1).fill = tot_fill
    ws2.cell(row=tr2, column=1).alignment = lft
    ws2.cell(row=tr2, column=1).border = bdr
    tot_map2 = {5:'v_serv',6:'iss_ret',7:'pis_ret',8:'cofins_ret',9:'ir_ret',
                10:'csll_ret',11:'inss_ret',12:'cbs_ret',13:'ibs_ret'}
    for c in range(2, len(cols2)+1):
        cell = ws2.cell(row=tr2, column=c); cell.fill = tot_fill; cell.border = bdr
        if c in tot_map2:
            cell.value = sum(r[tot_map2[c]] for r in ret_fed_rows)
            cell.number_format = money; cell.alignment = rgt; cell.font = tot_font
        elif c == 14:
            cell.value = sum(r['total_ret']+r['iss_ret'] for r in ret_rows)
            cell.number_format = money; cell.alignment = rgt; cell.font = tot_font
    ws2.column_dimensions['C'].width = 22; ws2.column_dimensions['D'].width = 40
    ws2.freeze_panes = 'A2'

    # Sheet 3: Retencao Municipal
    ws_mun = wb.create_sheet('Retencao Municipal')
    for c, h in enumerate(cols2, 1):
        cell = ws_mun.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = PatternFill('solid', start_color='1A7A4A'); cell.alignment = ctr; cell.border = bdr
    for i, r in enumerate(ret_mun_rows, start=2):
        fill = PatternFill('solid', start_color='E6F4ED') if i%2==0 else PatternFill('solid', start_color='FFFFFF')
        tot = r['iss_ret']
        vals = [r['numero'],r['emissao'],r['cnpj_emit'],r['nome_emit'],
                r['v_serv'],r['iss_ret'],r['pis_ret'],r['cofins_ret'],
                r['ir_ret'],r['csll_ret'],r['inss_ret'],tot]
        for c, v in enumerate(vals, 1):
            cell = ws_mun.cell(row=i, column=c, value=v)
            cell.font = nrm_font; cell.fill = fill; cell.border = bdr
            if c in money_cols2: cell.number_format = money; cell.alignment = rgt
            elif c in {1,2}: cell.alignment = ctr
            else: cell.alignment = lft
    tr_mun = len(ret_mun_rows)+2
    ws_mun.cell(row=tr_mun, column=1, value='TOTAIS').font = tot_font
    ws_mun.cell(row=tr_mun, column=1).fill = tot_fill
    ws_mun.cell(row=tr_mun, column=1).alignment = lft
    ws_mun.cell(row=tr_mun, column=1).border = bdr
    for c in range(2, len(cols2)+1):
        cell = ws_mun.cell(row=tr_mun, column=c); cell.fill = tot_fill; cell.border = bdr
        if c in tot_map2:
            cell.value = sum(r[tot_map2[c]] for r in ret_mun_rows)
            cell.number_format = money; cell.alignment = rgt; cell.font = tot_font
        elif c == 12:
            cell.value = sum(r['iss_ret'] for r in ret_mun_rows)
            cell.number_format = money; cell.alignment = rgt; cell.font = tot_font
    ws_mun.column_dimensions['C'].width = 22; ws_mun.column_dimensions['D'].width = 40
    ws_mun.freeze_panes = 'A2'

    ws3 = wb.create_sheet('Resumo por Servico')
    cols3 = ['Cod. Tributacao','Descr. Tributacao','Qtd. Notas','Vl. Total Servicos',
             'ISS Valor Total','ISS Ret. Total','Pis Ret. Total','Cofins Ret. Total',
             'IR Ret. Total','CSLL Ret. Total','INSS Ret. Total','Vl. Liquido Total']
    for c, h in enumerate(cols3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = ctr; cell.border = bdr
    groups = defaultdict(lambda: {'desc':'','count':0,'v_serv':0,'iss_val':0,'iss_ret':0,
                                   'pis':0,'cofins':0,'ir':0,'csll':0,'inss':0,'v_liq':0})
    for r in rows:
        k = r['cod_trib'] or 'N/A'; g = groups[k]
        g['desc'] = r['desc_trib']; g['count'] += 1
        g['v_serv'] += r['v_serv']; g['iss_val'] += r['iss_val']
        g['iss_ret'] += r['iss_ret']; g['pis'] += r['pis_ret']
        g['cofins'] += r['cofins_ret']; g['ir'] += r['ir_ret']
        g['csll'] += r['csll_ret']; g['inss'] += r['inss_ret']
        g['v_liq'] += r['v_liq']
    money_cols3 = {4,5,6,7,8,9,10,11,12}
    for i, (k, g) in enumerate(groups.items(), start=2):
        fill = PatternFill('solid', start_color='F0F5FC') if i%2==0 else PatternFill('solid', start_color='FFFFFF')
        vals = [k,g['desc'],g['count'],g['v_serv'],g['iss_val'],g['iss_ret'],
                g['pis'],g['cofins'],g['ir'],g['csll'],g['inss'],g['v_liq']]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=i, column=c, value=v)
            cell.font = nrm_font; cell.fill = fill; cell.border = bdr
            if c in money_cols3: cell.number_format = money; cell.alignment = rgt
            else: cell.alignment = lft
    tr3 = len(groups)+2
    ws3.cell(row=tr3, column=1, value='TOTAIS GERAIS').font = tot_font
    ws3.cell(row=tr3, column=1).fill = tot_fill
    ws3.cell(row=tr3, column=1).alignment = lft
    ws3.cell(row=tr3, column=1).border = bdr
    for c in range(2, len(cols3)+1):
        cell = ws3.cell(row=tr3, column=c); cell.fill = tot_fill; cell.border = bdr
    ws3.column_dimensions['B'].width = 60; ws3.freeze_panes = 'A2'

    ws4 = wb.create_sheet('Notas Canceladas')
    for c, h in enumerate(cols1, 1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = PatternFill('solid', start_color='C0392B')
        cell.alignment = ctr; cell.border = bdr
    for i, r in enumerate(cancel_rows, start=2):
        fill = PatternFill('solid', start_color='F0F5FC') if i%2==0 else PatternFill('solid', start_color='FFFFFF')
        vals = [r['numero'],r['local_prest'],r['emissao'],r['cnpj_emit'],r['nome_emit'],
                r['cnpj_toma'],r['nome_toma'],r['nbs'],r['desc_nbs'],r['cod_trib'],
                r['desc_trib'],r['desc_serv'],r['v_serv'],r['iss_aliq'],r['iss_val'],
                r['iss_ret'],r['pis_ret'],r['cofins_ret'],r['ir_ret'],r['csll_ret'],
                r['inss_ret'],r['cbs_ret'],r['ibs_ret'],r['v_liq'],r['obs']]
        for c, v in enumerate(vals, 1):
            cell = ws4.cell(row=i, column=c, value=v)
            cell.font = nrm_font; cell.fill = fill; cell.border = bdr
            cell.alignment = ctr if c in {1,2,3} else lft
    ws4.freeze_panes = 'A2'

    safe_name = company_name.replace('/','_').replace('\\','_').replace(':','_')
    out_path  = os.path.join(download_dir, f'Recebidas_NFS-e_{safe_name}_{month}.xlsx')
    wb.save(out_path)
    print(f'[OK] Planilha gerada: {out_path}', flush=True)
    return out_path

# ─────────────────────────────────────────────
# Page helpers
# ─────────────────────────────────────────────

def wait_for_page_ready(page, retries=10, timeout=120000):
    for attempt in range(retries):
        try:
            page.wait_for_load_state('networkidle', timeout=timeout)
            content = page.content()
            if '502' in content or 'Server Error' in content or 'service is unavailable' in content.lower():
                print(f'[AVISO] Portal erro servidor (tentativa {attempt+1}/{retries})', flush=True)
                time.sleep(10); page.reload(); continue
            return True
        except Exception as e:
            if attempt < retries - 1:
                print(f'[AVISO] Timeout pagina (tentativa {attempt+1}/{retries})', flush=True)
                time.sleep(5)
                try: page.reload()
                except: pass
            else: raise e
    return False

# ─────────────────────────────────────────────
# Scraping
# ─────────────────────────────────────────────

def scrape_page_rows(page):
    results = []
    rows = page.query_selector_all('tr[data-chave]')
    for row in rows:
        try:
            if 'nfse-cancelada' in (row.get_attribute('class') or ''):
                continue
            xml_link = row.query_selector("a[href*='/Download/NFSe/']")
            pdf_link = row.query_selector("a[href*='/Download/DANFSe/']")
            if xml_link and pdf_link:
                chave = xml_link.get_attribute('href').split('/')[-1]
                results.append({
                    'chave':   chave,
                    'xml_url': f'https://www.nfse.gov.br/emissornacional/DPS/ModalCaptcha/NFSe/{chave}',
                    'pdf_url': f'https://www.nfse.gov.br/emissornacional/DPS/ModalCaptcha/DANFSe/{chave}',
                })
        except:
            continue
    return results

def get_download_urls(page):
    results = []
    base_url = page.url
    pg = 1
    while True:
        page_rows = scrape_page_rows(page)
        results.extend(page_rows)
        print(f'[OK] Pagina {pg}: {len(page_rows)} notas', flush=True)
        proxima = page.query_selector("a[data-original-title='Próxima']")
        ultima  = page.query_selector("a[data-original-title='Última']")
        if not proxima and not ultima:
            break
        pg += 1
        next_url = re.sub(r'pg=\d+', f'pg={pg}', base_url) if 'pg=' in base_url else base_url + ('&' if '?' in base_url else '?') + f'pg={pg}'
        page.goto(next_url)
        wait_for_page_ready(page)
        if pg > 50:
            print('[AVISO] Safety stop at page 50', flush=True)
            break
    print(f'[OK] {len(results)} notas mapeadas em {pg} pagina(s)', flush=True)
    return results

# ─────────────────────────────────────────────
# Download via page.request (no captcha)
# ─────────────────────────────────────────────

def request_download(page, url, save_path, referer, retries=5):
    headers = {
        'Referer': referer,
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'Upgrade-Insecure-Requests': '1',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-User': '?1',
    }
    chave = url.split('/')[-1]
    for attempt in range(retries):
        try:
            response = page.request.get(url, headers=headers, timeout=120000)
            if response.status == 200:
                content = response.body()
                if len(content) < 100:
                    raise Exception(f'Resposta muito pequena ({len(content)} bytes)')
                if content[:5] in (b'<!DOC', b'<html', b'<HTML'):
                    raise Exception('Portal retornou HTML em vez do arquivo')
                with open(save_path, 'wb') as f:
                    f.write(content)
                return True
            elif response.status == 403:
                print(f'[AVISO] 403 em {chave[:20]} (tentativa {attempt+1}/{retries})', flush=True)
                time.sleep(5)
            elif response.status == 429:
                print(f'[AVISO] Rate limit 429, aguardando 30s...', flush=True)
                time.sleep(30)
            else:
                print(f'[AVISO] HTTP {response.status} (tentativa {attempt+1}/{retries})', flush=True)
                time.sleep(5)
        except Exception as e:
            if attempt < retries - 1:
                print(f'[AVISO] Falha download (tentativa {attempt+1}/{retries}): {str(e)[:80]}', flush=True)
                time.sleep(8)
            else:
                print(f'[ERRO] Falha ao baixar {chave[:20]}: {e}', flush=True)
                return False
    return False

# ─────────────────────────────────────────────
# Main reinf flow
# ─────────────────────────────────────────────

def download_files(page, download_urls, impostos_retidos, download_dir, company_name="", month=""):
    all_xml_dir = os.path.join(download_dir, "all", "xmls")
    all_pdf_dir = os.path.join(download_dir, "all", "pdfs")
    fed_xml_dir = os.path.join(download_dir, "federal", "xmls")
    fed_pdf_dir = os.path.join(download_dir, "federal", "pdfs")
    mun_xml_dir = os.path.join(download_dir, "municipal", "xmls")
    mun_pdf_dir = os.path.join(download_dir, "municipal", "pdfs")
    temp_dir    = os.path.join(download_dir, "temp")
    for d in [all_xml_dir, all_pdf_dir, fed_xml_dir, fed_pdf_dir, mun_xml_dir, mun_pdf_dir, temp_dir]:
        os.makedirs(d, exist_ok=True)

    referer    = page.url
    all_parsed = []
    total      = len(download_urls)
    downloaded = failed = federal_count = municipal_count = none_count = 0

    for i, url_info in enumerate(download_urls, 1):
        chave    = url_info["chave"]
        xml_path = os.path.join(all_xml_dir, f"{chave}.xml")
        pdf_path = os.path.join(all_pdf_dir, f"{chave}.pdf")
        print(f"[{i}/{total}] {chave[:20]}...", flush=True)

        xml_ok = request_download(page, url_info["xml_url"], xml_path, referer)
        if not xml_ok:
            failed += 1
            continue

        pdf_ok = request_download(page, url_info["pdf_url"], pdf_path, referer)
        if not pdf_ok:
            failed += 1

        data = parse_xml_full(xml_path)
        if data:
            all_parsed.append(data)
            nnfse = data["numero"]
            if data["is_federal"]:
                shutil.copy2(xml_path, os.path.join(fed_xml_dir, f"{chave}.xml"))
                if pdf_ok: shutil.copy2(pdf_path, os.path.join(fed_pdf_dir, f"{chave}.pdf"))
                federal_count += 1
                print(f"[OK] FEDERAL | Nota {nnfse}", flush=True)
            elif data["is_municipal"]:
                shutil.copy2(xml_path, os.path.join(mun_xml_dir, f"{chave}.xml"))
                if pdf_ok: shutil.copy2(pdf_path, os.path.join(mun_pdf_dir, f"{chave}.pdf"))
                municipal_count += 1
                print(f"[OK] MUNICIPAL | Nota {nnfse}", flush=True)
            else:
                none_count += 1
            downloaded += 1
        time.sleep(0.2)

    shutil.rmtree(temp_dir, ignore_errors=True)

    if company_name and month:
        try:
            generate_recebidas_excel(all_parsed, company_name, month, download_dir)
        except Exception as e:
            print(f"[AVISO] Erro ao gerar planilha: {e}", flush=True)

    print(f"[OK] {downloaded} notas - {federal_count} federal, {municipal_count} municipal, {none_count} sem retencao", flush=True)
    if failed > 0:
        print(f"[AVISO] {failed} falha(s)", flush=True)


def download_files_all(page, download_dir):
    notas_dir = os.path.join(download_dir, 'notas')
    os.makedirs(notas_dir, exist_ok=True)
    print('[INFO] Modo completo - baixando todas as notas', flush=True)
    all_urls = get_download_urls(page)
    if not all_urls:
        print('[AVISO] Nenhuma nota encontrada', flush=True)
        return None
    referer = page.url
    downloaded = failed = 0
    total = len(all_urls)
    for i, url_info in enumerate(all_urls, 1):
        chave = url_info['chave']
        print(f'[{i}/{total}] Nota {chave[:20]}...', flush=True)
        xml_ok = request_download(page, url_info['xml_url'], os.path.join(notas_dir, f'{chave}.xml'), referer)
        pdf_ok = request_download(page, url_info['pdf_url'], os.path.join(notas_dir, f'{chave}.pdf'), referer)
        if xml_ok or pdf_ok: downloaded += 1
        else: failed += 1
        time.sleep(0.3)
    print(f'[OK] {downloaded} notas baixadas | {failed} falha(s)', flush=True)
    return notas_dir if downloaded > 0 else None

# ─────────────────────────────────────────────
# Emitidas flow
# ─────────────────────────────────────────────

def generate_emitidas_excel(rows, company_name, month, download_dir):
    wb = Workbook()
    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill = PatternFill('solid', start_color='1A7A4A')
    tot_font = Font(name='Arial', bold=True, size=10)
    tot_fill = PatternFill('solid', start_color='D5F0E0')
    nrm_font = Font(name='Arial', size=10)
    ttl_font = Font(name='Arial', bold=True, size=11, color='1A7A4A')
    ctr  = Alignment(horizontal='center', vertical='center')
    lft  = Alignment(horizontal='left',   vertical='center')
    rgt  = Alignment(horizontal='right',  vertical='center')
    thin = Side(style='thin', color='BFBFBF')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    money = '#,##0.00'

    cancel_rows = [r for r in rows if r['cancelada']]

    cols = ['Nr NFSe', 'Emissao', 'CNPJ/CPF Tomador', 'Razao Tomador',
            'Cod. Tributacao', 'Descr. Tributacao', 'Descr. Servico',
            'Vl. Servico', 'ISS Valor', 'ISS Ret.',
            'Pis Ret.', 'Cofins Ret.', 'IR Ret.', 'CSLL Ret.', 'INSS Ret.',
            'CBS Ret.', 'IBS Ret.', 'Vl. Liquido', 'Situacao']
    money_cols = {8, 9, 10, 11, 12, 13, 14, 15, 16, 17}

    ws = wb.active
    ws.title = 'Notas Emitidas'
    ws.merge_cells(f'A1:S1')
    ws['A1'] = company_name
    ws['A1'].font = ttl_font
    ws['A1'].alignment = lft
    ws.merge_cells('A2:S2')
    ws['A2'] = f"Relatorio de Notas Emitidas   Gerado em: {datetime.now().strftime('%d/%m/%Y as %H:%M:%S')}"
    ws['A2'].font = Font(name='Arial', size=9, color='7F8C8D')
    for c, h in enumerate(cols, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = ctr; cell.border = bdr
    ws.row_dimensions[3].height = 16

    active_rows = [r for r in rows if not r['cancelada']]
    for i, r in enumerate(active_rows, start=4):
        fill = PatternFill('solid', start_color='F0FAF4') if i % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
        situacao = 'ATIVA'
        vals = [r['numero'], r['emissao'], r['cnpj_toma'], r['nome_toma'],
                r['cod_trib'], r['desc_trib'], r['desc_serv'],
                r['v_serv'], r['iss_val'], r['iss_ret'],
                r['pis_ret'], r['cofins_ret'], r['ir_ret'], r['csll_ret'], r['inss_ret'],
                r['cbs_ret'], r['ibs_ret'], r['v_liq'], situacao]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = nrm_font; cell.fill = fill; cell.border = bdr
            if c in money_cols: cell.number_format = money; cell.alignment = rgt
            elif c in {1, 2}: cell.alignment = ctr
            else: cell.alignment = lft

    tr = len(active_rows) + 4
    ws.cell(row=tr, column=1, value='TOTAIS').font = tot_font
    ws.cell(row=tr, column=1).fill = tot_fill
    ws.cell(row=tr, column=1).alignment = lft
    ws.cell(row=tr, column=1).border = bdr
    tot_map = {8: 'v_serv', 9: 'iss_val', 10: 'iss_ret', 11: 'pis_ret',
               12: 'cofins_ret', 13: 'ir_ret', 14: 'csll_ret', 15: 'inss_ret',
               16: 'cbs_ret', 17: 'ibs_ret', 18: 'v_liq'}
    for c in range(2, len(cols) + 1):
        cell = ws.cell(row=tr, column=c); cell.fill = tot_fill; cell.border = bdr
        if c in tot_map:
            cell.value = sum(r[tot_map[c]] for r in active_rows)
            cell.number_format = money; cell.alignment = rgt; cell.font = tot_font

    # Canceladas sheet
    ws_cancel = wb.create_sheet('Notas Canceladas')
    for c, h in enumerate(cols, 1):
        cell = ws_cancel.cell(row=1, column=c, value=h)
        cell.font = hdr_font; cell.fill = PatternFill('solid', start_color='C0392B')
        cell.alignment = ctr; cell.border = bdr
    for i, r in enumerate(cancel_rows, start=2):
        fill = PatternFill('solid', start_color='FDF0EF') if i % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
        vals = [r['numero'], r['emissao'], r['cnpj_toma'], r['nome_toma'],
                r['cod_trib'], r['desc_trib'], r['desc_serv'],
                r['v_serv'], r['iss_val'], r['iss_ret'],
                r['pis_ret'], r['cofins_ret'], r['ir_ret'], r['csll_ret'], r['inss_ret'],
                r['cbs_ret'], r['ibs_ret'], r['v_liq'], 'CANCELADA']
        for c, v in enumerate(vals, 1):
            cell = ws_cancel.cell(row=i, column=c, value=v)
            cell.font = nrm_font; cell.fill = fill; cell.border = bdr
            if c in money_cols: cell.number_format = money; cell.alignment = rgt
            elif c in {1, 2}: cell.alignment = ctr
            else: cell.alignment = lft
    ws_cancel.freeze_panes = 'A2'

    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 50
    ws.freeze_panes = 'A4'

    safe_name = company_name.replace('/', '_').replace('\\', '_').replace(':', '_')
    out_path = os.path.join(download_dir, f'Emitidas_NFS-e_{safe_name}_{month}.xlsx')
    wb.save(out_path)
    print(f'[OK] Planilha emitidas gerada: {out_path}', flush=True)
    return out_path


def download_files_emitidas(page, download_dir, company_name="", month=""):
    emitidas_xml_dir = os.path.join(download_dir, 'emitidas', 'xmls')
    emitidas_pdf_dir = os.path.join(download_dir, 'emitidas', 'pdfs')
    for d in [emitidas_xml_dir, emitidas_pdf_dir]:
        os.makedirs(d, exist_ok=True)

    print('[INFO] Modo emitidas - mapeando notas...', flush=True)
    all_urls = get_download_urls(page)
    if not all_urls:
        print('[AVISO] Nenhuma nota emitida encontrada', flush=True)
        if company_name and month:
            generate_emitidas_excel([], company_name, month, download_dir)
        return None

    referer    = page.url
    all_parsed = []
    downloaded = failed = 0
    total      = len(all_urls)

    for i, url_info in enumerate(all_urls, 1):
        chave    = url_info['chave']
        xml_path = os.path.join(emitidas_xml_dir, f'{chave}.xml')
        pdf_path = os.path.join(emitidas_pdf_dir, f'{chave}.pdf')
        print(f'[{i}/{total}] {chave[:20]}...', flush=True)

        xml_ok = request_download(page, url_info['xml_url'], xml_path, referer)
        pdf_ok = request_download(page, url_info['pdf_url'], pdf_path, referer)

        if xml_ok:
            data = parse_xml_full(xml_path)
            if data:
                all_parsed.append(data)
            downloaded += 1
        else:
            failed += 1
        time.sleep(0.3)

    if company_name and month:
        try:
            generate_emitidas_excel(all_parsed, company_name, month, download_dir)
        except Exception as e:
            print(f'[AVISO] Erro ao gerar planilha emitidas: {e}', flush=True)

    print(f'[OK] {downloaded} notas emitidas baixadas | {failed} falha(s)', flush=True)
    return emitidas_xml_dir if downloaded > 0 else None
