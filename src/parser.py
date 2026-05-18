import openpyxl

def parse_impostos_retidos(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # find the Impostos Retidos sheet
    sheet = None
    for name in wb.sheetnames:
        if "retido" in name.lower():
            sheet = wb[name]
            break

    if not sheet:
        print("[AVISO] Aba 'Impostos Retidos' nao encontrada")
        return []

    notes = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # skip empty rows and TOTAIS row
        if row[0] is None or str(row[0]).upper() == "TOTAIS":
            continue
        notes.append({
            "numero":         str(row[0]).strip(),
            "emissao":        str(row[1]).strip() if row[1] else "",
            "cnpj_emitente":  str(row[2]).strip() if row[2] else "",
            "razao_emitente": str(row[3]).strip() if row[3] else "",
            "vl_servico":     row[4] or 0,
            "iss_ret":        row[5] or 0,
            "pis_ret":        row[6] or 0,
            "cofins_ret":     row[7] or 0,
            "ir_ret":         row[8] or 0,
            "csll_ret":       row[9] or 0,
            "inss_ret":       row[10] or 0,
            "total_retido":   row[11] or 0,
        })

    print(f"[OK] {len(notes)} nota(s) com impostos retidos")
    return notes