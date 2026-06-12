"""
generate_visualizar_excel.py
Generates master Excel from scraped Visualizar data.
3 tabs: Todas as Notas | RetenÃ§Ã£o Federal | RetenÃ§Ã£o Municipal
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def _make_styles():
    header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill_blue  = PatternFill('solid', start_color='1A56A0')
    header_fill_red   = PatternFill('solid', start_color='C0392B')
    header_fill_green = PatternFill('solid', start_color='1A7A4A')
    title_font   = Font(name='Arial', bold=True, size=11)
    total_font   = Font(name='Arial', bold=True, size=10)
    total_fill   = PatternFill('solid', start_color='F9E79F')
    normal_font  = Font(name='Arial', size=10)
    center       = Alignment(horizontal='center', vertical='center')
    left         = Alignment(horizontal='left',   vertical='center')
    right        = Alignment(horizontal='right',  vertical='center')
    thin         = Side(style='thin', color='BFBFBF')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        'header_font': header_font,
        'header_fill_blue': header_fill_blue,
        'header_fill_red': header_fill_red,
        'header_fill_green': header_fill_green,
        'title_font': title_font,
        'total_font': total_font,
        'total_fill': total_fill,
        'normal_font': normal_font,
        'center': center,
        'left': left,
        'right': right,
        'border': border,
    }

def _write_header(ws, row, headers, s, fill):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = s['header_font']
        cell.fill      = fill
        cell.alignment = s['center']
        cell.border    = s['border']
    ws.row_dimensions[row].height = 16

def _write_row(ws, row_idx, values, s, money_cols=None, pct_cols=None):
    fill = PatternFill('solid', start_color='F0F5FC') if row_idx % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
    money_cols = money_cols or []
    pct_cols   = pct_cols   or []
    for col, val in enumerate(values, 1):
        cell        = ws.cell(row=row_idx, column=col, value=val)
        cell.font   = s['normal_font']
        cell.fill   = fill
        cell.border = s['border']
        if col in money_cols:
            cell.number_format = '#,##0.00'
            cell.alignment     = s['right']
        elif col in pct_cols:
            cell.number_format = '0.00'
            cell.alignment     = s['center']
        elif col == 1:
            cell.alignment = s['center']
        else:
            cell.alignment = s['left']

def _write_totals(ws, row, col_count, total_dict, s):
    """Write totals row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = s['total_fill']
        cell.border = s['border']
        cell.font   = s['total_font']
    ws.cell(row=row, column=1, value='TOTAL').alignment = s['left']
    for col, val in total_dict.items():
        cell = ws.cell(row=row, column=col, value=val)
        cell.number_format = '#,##0.00'
        cell.alignment     = s['right']

def generate_visualizar_excel(company_name, month, notas, out_dir):
    """
    notas: list of dicts from scrape_visualizar()
    """
    if not notas:
        return None

    os.makedirs(out_dir, exist_ok=True)
    safe_name = company_name.replace('/', '_').replace('\\', '_').replace(':', '_')
    out_path  = os.path.join(out_dir, f'NFS-e_{safe_name}_{month}.xlsx')

    wb = Workbook()
    s  = _make_styles()
    gen_time = datetime.now().strftime('%d/%m/%Y %H:%M')

    # â”€â”€ Tab 1: Todas as Notas â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws1 = wb.active
    ws1.title = 'Todas as Notas'

    ws1.merge_cells('A1:T1')
    ws1['A1']           = f'NFS-e â€” {company_name} â€” {month} â€” Gerado em {gen_time}'
    ws1['A1'].font      = s['title_font']
    ws1['A1'].alignment = s['left']
    ws1.row_dimensions[1].height = 20

    headers_all = [
        'NÂº DPS', 'Data EmissÃ£o', 'SituaÃ§Ã£o',
        'Emitente CNPJ', 'Emitente Nome',
        'Tomador CNPJ', 'Tomador Nome',
        'MunicÃ­pio', 'Vl. ServiÃ§o', 'Desconto', 'Base ISS',
        'AlÃ­q ISS %', 'Vl. ISS', 'Ret. ISS',
        'Vl. PIS', 'Vl. COFINS', 'Vl. IR', 'Vl. CSLL', 'Vl. INSS',
        'ClassificaÃ§Ã£o'
    ]
    _write_header(ws1, 2, headers_all, s, s['header_fill_blue'])

    money_cols_all = {9, 10, 11, 13, 15, 16, 17, 18, 19}
    pct_cols_all   = {12}

    for i, n in enumerate(notas, start=3):
        if n.get('is_cancelada'):
            classif = 'CANCELADA'
        else:
            labels = []
            if n['is_federal']:
                labels.append('FEDERAL')
            if n['is_municipal']:
                labels.append('MUNICIPAL')
            classif = ' + '.join(labels) if labels else 'SEM RETENÃ‡ÃƒO'
        vals = [
            n['numero'], n['data_emissao'], n['situacao'],
            n['emit_cnpj'], n['emit_nome'],
            n['toma_cnpj'], n['toma_nome'],
            n['mun_incidencia'],
            n['v_servico'], n['desconto'], n['base_calculo'],
            n['aliquota_iss'], n['v_issqn'], n['ret_issqn'],
            n['v_pis'], n['v_cofins'], n['v_irrf'], n['v_csll'], n['v_inss'],
            classif
        ]
        _write_row(ws1, i, vals, s, money_cols=money_cols_all, pct_cols=pct_cols_all)

    total_row = len(notas) + 3
    _write_totals(ws1, total_row, len(headers_all), {
        9:  sum(n['v_servico'] for n in notas),
        13: sum(n['v_issqn']   for n in notas),
        15: sum(n['v_pis']     for n in notas),
        16: sum(n['v_cofins']  for n in notas),
        17: sum(n['v_irrf']    for n in notas),
        18: sum(n['v_csll']    for n in notas),
        19: sum(n['v_inss']    for n in notas),
    }, s)

    col_widths_all = [10, 18, 20, 22, 40, 22, 40, 20, 14, 12, 12, 10, 12, 20, 12, 12, 12, 12, 12, 15]
    for i, w in enumerate(col_widths_all, 1):
        ws1.column_dimensions[ws1.cell(row=2, column=i).column_letter].width = w
    ws1.freeze_panes = 'A3'

    # â”€â”€ Tab 2: RetenÃ§Ã£o Federal â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws2 = wb.create_sheet('RetenÃ§Ã£o Federal')
    fed  = [n for n in notas if n['is_federal'] and not n.get('is_cancelada')]

    ws2.merge_cells('A1:N1')
    ws2['A1']           = f'NFS-e COM RETENÃ‡ÃƒO FEDERAL â€” {company_name} â€” {month}'
    ws2['A1'].font      = Font(name='Arial', bold=True, size=11, color='C0392B')
    ws2['A1'].alignment = s['left']
    ws2.row_dimensions[1].height = 20

    headers_fed = [
        'NÂº DPS', 'Data EmissÃ£o',
        'Emitente CNPJ', 'Emitente Nome',
        'Tomador CNPJ', 'Tomador Nome',
        'Vl. ServiÃ§o', 'Vl. PIS', 'Vl. COFINS',
        'Vl. IR', 'Vl. CSLL', 'Vl. INSS', 'Total Retido',
        'Sit. PIS/COFINS'
    ]
    _write_header(ws2, 2, headers_fed, s, s['header_fill_red'])

    money_cols_fed = {7, 8, 9, 10, 11, 12, 13}

    for i, n in enumerate(fed, start=3):
        total_ret = n['v_pis'] + n['v_cofins'] + n['v_irrf'] + n['v_csll'] + n['v_inss']
        vals = [
            n['numero'], n['data_emissao'],
            n['emit_cnpj'], n['emit_nome'],
            n['toma_cnpj'], n['toma_nome'],
            n['v_servico'], n['v_pis'], n['v_cofins'],
            n['v_irrf'], n['v_csll'], n['v_inss'], total_ret,
            n['sit_pis_cofins']
        ]
        _write_row(ws2, i, vals, s, money_cols=money_cols_fed)

    if fed:
        total_row2 = len(fed) + 3
        _write_totals(ws2, total_row2, len(headers_fed), {
            7:  sum(n['v_servico'] for n in fed),
            8:  sum(n['v_pis']     for n in fed),
            9:  sum(n['v_cofins']  for n in fed),
            10: sum(n['v_irrf']    for n in fed),
            11: sum(n['v_csll']    for n in fed),
            12: sum(n['v_inss']    for n in fed),
            13: sum(n['v_pis'] + n['v_cofins'] + n['v_irrf'] + n['v_csll'] + n['v_inss'] for n in fed),
        }, s)

    col_widths_fed = [10, 18, 22, 40, 22, 40, 14, 12, 12, 12, 12, 12, 14, 35]
    for i, w in enumerate(col_widths_fed, 1):
        ws2.column_dimensions[ws2.cell(row=2, column=i).column_letter].width = w
    ws2.freeze_panes = 'A3'

    # â”€â”€ Tab 3: RetenÃ§Ã£o Municipal â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws3 = wb.create_sheet('RetenÃ§Ã£o Municipal')
    mun  = [n for n in notas if n['is_municipal'] and not n.get('is_cancelada')]

    ws3.merge_cells('A1:K1')
    ws3['A1']           = f'NFS-e COM RETENÃ‡ÃƒO MUNICIPAL â€” {company_name} â€” {month}'
    ws3['A1'].font      = Font(name='Arial', bold=True, size=11, color='1A7A4A')
    ws3['A1'].alignment = s['left']
    ws3.row_dimensions[1].height = 20

    headers_mun = [
        'NÂº DPS', 'Data EmissÃ£o',
        'Emitente CNPJ', 'Emitente Nome',
        'Tomador CNPJ', 'Tomador Nome',
        'MunicÃ­pio', 'Vl. ServiÃ§o', 'Base ISS',
        'AlÃ­q ISS %', 'Vl. ISS Retido'
    ]
    _write_header(ws3, 2, headers_mun, s, s['header_fill_green'])

    money_cols_mun = {8, 9, 11}
    pct_cols_mun   = {10}

    for i, n in enumerate(mun, start=3):
        vals = [
            n['numero'], n['data_emissao'],
            n['emit_cnpj'], n['emit_nome'],
            n['toma_cnpj'], n['toma_nome'],
            n['mun_incidencia'],
            n['v_servico'], n['base_calculo'],
            n['aliquota_iss'], n['v_issqn']
        ]
        _write_row(ws3, i, vals, s, money_cols=money_cols_mun, pct_cols=pct_cols_mun)

    if mun:
        total_row3 = len(mun) + 3
        _write_totals(ws3, total_row3, len(headers_mun), {
            8:  sum(n['v_servico']   for n in mun),
            9:  sum(n['base_calculo']for n in mun),
            11: sum(n['v_issqn']     for n in mun),
        }, s)

    col_widths_mun = [10, 18, 22, 40, 22, 40, 20, 14, 12, 10, 14]
    for i, w in enumerate(col_widths_mun, 1):
        ws3.column_dimensions[ws3.cell(row=2, column=i).column_letter].width = w
    ws3.freeze_panes = 'A3'

    # â”€â”€ Save (with fallback if file is locked/open elsewhere) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    try:
        wb.save(out_path)
        print(f'[OK] Excel salvo: {out_path}', flush=True)
        return out_path
    except PermissionError:
        run_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        alt_path = os.path.join(out_dir, f'NFS-e_{safe_name}_{month}_{run_id}.xlsx')
        try:
            wb.save(alt_path)
            print(f'[AVISO] Arquivo original estava aberto/bloqueado â€” salvo como: {alt_path}', flush=True)
            return alt_path
        except Exception as e:
            print(f'[ERRO] Falha ao salvar Excel mesmo com nome alternativo: {e}', flush=True)
            raise
