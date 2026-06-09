import os, re, uuid, threading, zipfile, io, time, requests, logging, json, sqlite3
import pandas as pd
import pdfplumber
from datetime import timedelta, datetime
from flask import Flask, request, jsonify, send_file, render_template
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def load_config():
    path = os.path.join(BASE_DIR, 'config.json')
    defaults = {
        "codigo_fornecedor": "148",
        "codigo_folha": "6575",
        "codigo_nao_identificado": "5",
        "port": 5002,
        "comprovante_tolerance_days": 1,
        "receita_ws_enabled": True,
        "receita_ws_rpm": 3
    }
    if os.path.exists(path):
        with open(path) as f:
            data = json.load(f)
        defaults.update(data)
    return defaults

def save_config(cfg):
    with open(os.path.join(BASE_DIR, 'config.json'), 'w') as f:
        json.dump(cfg, f, indent=4)

LOG_DIR = os.path.join(BASE_DIR, 'logs')
os.makedirs(LOG_DIR, exist_ok=True)
log_file = os.path.join(LOG_DIR, datetime.now().strftime('%Y-%m-%d') + '.log')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('fam')

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
app.config['OUTPUT_FOLDER'] = os.path.join(BASE_DIR, 'outputs')
app.config['DATA_FOLDER']   = os.path.join(BASE_DIR, 'data')

for folder in [app.config['UPLOAD_FOLDER'], app.config['OUTPUT_FOLDER'], app.config['DATA_FOLDER']]:
    os.makedirs(folder, exist_ok=True)

tasks = {}
DB_PATH = os.path.join(BASE_DIR, 'data', 'fam.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS fornecedores (
                cnpj TEXT PRIMARY KEY,
                nome TEXT NOT NULL,
                source TEXT DEFAULT 'CADASTRO',
                confirmed INTEGER DEFAULT 1,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS funcionarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nome TEXT UNIQUE NOT NULL,
                confirmed INTEGER DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS cheques (
                numero TEXT PRIMARY KEY,
                portador TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS cnpj_cache (
                cnpj TEXT PRIMARY KEY,
                nome TEXT,
                looked_up_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS run_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_date TEXT DEFAULT (datetime('now')),
                total_rows INTEGER,
                ok_count INTEGER,
                revisar_count INTEGER,
                nao_encontrado_count INTEGER,
                novos_fornecedores INTEGER
            );
            CREATE TABLE IF NOT EXISTS corrections (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                complemento_text TEXT,
                cnpj_identified TEXT,
                correct_participant TEXT,
                correct_codigo TEXT,
                corrected_at TEXT DEFAULT (datetime('now'))
            );
        ''')
    logger.info("DB initialized")

def migrate_csvs_to_db():
    """Migrate CSVs to DB only if DB tables are empty — runs once."""
    data_dir = app.config['DATA_FOLDER']
    with get_db() as conn:
        forn_count = conn.execute('SELECT COUNT(*) FROM fornecedores').fetchone()[0]
        forn_path = os.path.join(data_dir, 'fornecedores.csv')
        if forn_count == 0 and os.path.exists(forn_path):
            df = pd.read_csv(forn_path, dtype=str).fillna('')
            count = 0
            for _, row in df.iterrows():
                cnpj = re.sub(r'[^\d]', '', row.get('CNPJ', ''))
                nome = row.get('NOME', '').strip()
                if cnpj and nome:
                    conn.execute('INSERT OR IGNORE INTO fornecedores (cnpj, nome, source) VALUES (?,?,?)', (cnpj, nome, 'CADASTRO'))
                    count += 1
            logger.info(f"Migrated {count} fornecedores")

        func_count = conn.execute('SELECT COUNT(*) FROM funcionarios').fetchone()[0]
        func_path = os.path.join(data_dir, 'funcionarios.csv')
        if func_count == 0 and os.path.exists(func_path):
            df = pd.read_csv(func_path, dtype=str).fillna('')
            count = 0
            for _, row in df.iterrows():
                nome = row.get('NOME', '').strip()
                if nome:
                    conn.execute('INSERT OR IGNORE INTO funcionarios (nome) VALUES (?)', (nome,))
                    count += 1
            logger.info(f"Migrated {count} funcionarios")

        cheq_count = conn.execute('SELECT COUNT(*) FROM cheques').fetchone()[0]
        cheq_path = os.path.join(data_dir, 'cheques.csv')
        if cheq_count == 0 and os.path.exists(cheq_path):
            df = pd.read_csv(cheq_path, dtype=str).fillna('')
            count = 0
            for _, row in df.iterrows():
                num = row.get('Numero', '').strip()
                por = row.get('Portador', '').strip()
                if num and por:
                    conn.execute('INSERT OR IGNORE INTO cheques (numero, portador) VALUES (?,?)', (num, por))
                    count += 1
            logger.info(f"Migrated {count} cheques")

def build_maps():
    with get_db() as conn:
        forn_rows  = conn.execute('SELECT cnpj, nome FROM fornecedores WHERE confirmed=1').fetchall()
        func_rows  = conn.execute('SELECT nome FROM funcionarios WHERE confirmed=1').fetchall()
        cheq_rows  = conn.execute('SELECT numero, portador FROM cheques').fetchall()
        cache_rows = conn.execute('SELECT cnpj, nome FROM cnpj_cache').fetchall()
    forn_map    = {row['cnpj']: row['nome'] for row in forn_rows}
    func_names  = [row['nome'].upper() for row in func_rows]
    cheques_map = {row['numero']: row['portador'] for row in cheq_rows}
    cnpj_cache  = {row['cnpj']: row['nome'] for row in cache_rows}
    return forn_map, func_names, cheques_map, cnpj_cache

# ── CNPJ / CPF extraction ────────────────────────────────────────────────────

CNPJ_RE = re.compile(r'\b(\d{2}[\.\-]?\d{3}[\.\-]?\d{3}[\/]?\d{4}[\-]?\d{2}|\d{14})\b')
CPF_RE  = re.compile(r'\b\d{3}[\.\-]?\d{3}[\.\-]?\d{3}[\-]?\d{2}\b')
FAM_OWN_CNPJ = '04957294000103'   # FAM Metal's own CNPJ — self-transfers

def clean_cnpj(raw): return re.sub(r'[^\d]', '', raw)

def extract_cnpj(text):
    text = str(text)
    m = CNPJ_RE.search(text)
    if m:
        c = clean_cnpj(m.group())
        return c
    m = re.search(r'(?<!\d)(\d{14})(?!\d)', text)
    if m:
        digits = m.group(1)
        if digits[:2] != '00':
            return digits
    return None

def extract_cpf(text):
    m = CPF_RE.search(str(text))
    return re.sub(r'[^\d]', '', m.group()) if m else None

def is_self_transfer(text):
    """Returns True if the transaction is FAM paying itself."""
    return FAM_OWN_CNPJ in re.sub(r'[^\d]', '', str(text))

# ── Name extraction — covers ALL real SCI complemento patterns ───────────────

def extract_name_from_complemento(text):
    """
    Extract supplier/payee name from Complemento field.
    Returns (name: str | None, origin_hint: str | None)
    origin_hint is one of: 'BOLETO_109', 'BOLETO_144_PIX', 'BOLETO_PREF',
                            'PIX_SICREDI', 'TRANSF', 'TED', 'DEBITO_TED',
                            'DEBITO_CONV', 'CARTAO', 'LIQUIDACAO', None
    """
    text = str(text).strip()
    if not text:
        return None, None

    # ── patterns ordered most-specific → least-specific ─────────────────────

    # 1. "109 Pagamento de Boleto NAME - ref"  (419 rows, name-only, no CNPJ)
    m = re.match(
        r'^109\s+Pagamento\s+de\s+Boleto\s+(.+?)\s*-\s*[\d\.]+\s*$',
        text, re.IGNORECASE)
    if m: return _clean(m.group(1)), 'BOLETO_109'

    # 2. "144 Pix - Enviado NAME - ref"
    m = re.match(
        r'^144\s+Pix\s+-\s+Enviado\s+(.+?)\s*-\s*[\d\.]+\s*$',
        text, re.IGNORECASE)
    if m: return _clean(m.group(1)), 'BOLETO_144_PIX'

    # 3. "Pagamento de Boleto NAME - ref"  (mixed case, no numeric prefix)
    m = re.match(
        r'^Pagamento\s+de\s+Boleto\s+(.+?)\s*-\s*[\d\.]+\s*$',
        text, re.IGNORECASE)
    if m: return _clean(m.group(1)), 'BOLETO_PREF'

    # 4. "Pix - Enviado NAME - ref"
    m = re.match(
        r'^Pix\s+-\s+Enviado\s+(.+?)\s*-\s*[\d\.]+\s*$',
        text, re.IGNORECASE)
    if m: return _clean(m.group(1)), 'BOLETO_PREF'

    # 5. "470 Transferência enviada NAME - ref" or "Transferência enviada NAME - ref"
    m = re.match(
        r'^(?:470\s+)?Transfer[eê]ncia\s+enviada\s+(.+?)\s*-\s*[\d\.]+(?:\.[\d]+)*\s*$',
        text, re.IGNORECASE)
    if m: return _clean(m.group(1)), 'TRANSF'

    # 6. "TRANSF ENTRE CONTAS CNPJ NAME - ref"
    m = re.match(
        r'^TRANSF\s+ENTRE\s+CONTAS\s+\d+\s+(.+?)\s*-\s*\w+\d+\s*$',
        text, re.IGNORECASE)
    if m: return _clean(m.group(1)), 'TRANSF'

    # 7. "TED Transf.Eletr.Disponiv BANK AGENCIA CNPJ NAME - - ref"
    m = re.match(
        r'^TED\s+Transf[\w\.]+\s+\d+\s+\d+\s+\d+\s+(.+?)\s*-\s*-\s*[\d\.]+\s*$',
        text, re.IGNORECASE)
    if m: return _clean(m.group(1)), 'TED'

    # 8. "DEBITO TED/IB CNPJ NAME - ref"
    m = re.match(
        r'^DEBITO\s+TED/IB\s+\d+\s+(.+?)\s*-\s*\w+\d+\s*$',
        text, re.IGNORECASE)
    if m: return _clean(m.group(1)), 'DEBITO_TED'

    # 9. "DEBITO CONVENIOS CNPJ NAME - NAME"
    m = re.match(
        r'^DEBITO\s+CONVENIOS\s+\d+\s+(.+?)\s*-\s*.+$',
        text, re.IGNORECASE)
    if m: return _clean(m.group(1)), 'DEBITO_CONV'

    # 10. "Compra com Cartão NAME - ref"
    m = re.match(
        r'^Compra\s+com\s+Cart[aã]o\s+(.+?)\s*-\s*[\d\.]+\s*$',
        text, re.IGNORECASE)
    if m: return _clean(m.group(1)), 'CARTAO'

    # 11. "PAGAMENTO PIX [SICREDI] CNPJ/CPF NAME - suffix"
    #     Name comes AFTER the CNPJ/CPF token
    m = re.match(
        r'^PAGAMENTO\s+PIX(?:\s+SICREDI)?\s+[\d\.\/\-]+\s+(.+?)\s*-\s*(?:PIX_DEB|CX\w+|I\d+)\s*$',
        text, re.IGNORECASE)
    if m: return _clean(m.group(1)), 'PIX_SICREDI'

    # 12. "LIQUIDACAO BOLETO [SICREDI] CNPJ NAME -"
    m = re.match(
        r'^LIQUIDACAO\s+BOLETO(?:\s+SICREDI)?\s+[\d\.\/\-]+\s+(.+?)\s*-\s*.*$',
        text, re.IGNORECASE)
    if m:
        name = _clean(m.group(1))
        if name:
            # Strip parenthetical aliases like "(quero quero)"
            name = re.sub(r'\s*\(.*?\)', '', name).strip()
        if name:
            return name, 'LIQUIDACAO'

    return None, None


def _clean(name):
    """Strip trailing punctuation/spaces, uppercase, min length 2."""
    name = str(name).strip().rstrip(' -').strip()
    name = re.sub(r'\s+', ' ', name)
    if len(name) < 2:
        return None
    return name.upper()


def extract_pix_name(complemento):
    """Legacy — kept for employee folha matching. Returns raw name after PIX CNPJ/CPF."""
    if not complemento: return None
    m = re.search(
        r'(?:PAGAMENTO\s+)?PIX(?:\s+SICREDI)?\s+[\d.\/\-]+\s+(.+?)(?:\s+-\s*(?:PIX_|CX|I00).*)?$',
        str(complemento).strip(), re.IGNORECASE
    )
    return m.group(1).strip().rstrip('- ').upper() if m else None

def match_employee_name(pix_name, folha_names):
    pix_words = list(pix_name.upper().split())
    pix_set   = set(pix_words)
    for emp in folha_names:
        emp_words = list(emp.upper().split())
        emp_set   = set(emp_words)
        if len(pix_set & emp_set) >= 2: return True, pix_name[:40]
        if len(pix_words) >= 2 and len(emp_words) >= 2:
            if pix_words[0] == emp_words[0] and pix_words[-1] == emp_words[-1]:
                return True, pix_name[:40]
    return False, None

def extract_cheque_number(complemento):
    if not complemento: return None
    m = re.search(r'ALW0*(\d{4,})', str(complemento))
    if m: return m.group(1).lstrip('0') or '0'
    m = re.search(r'(\d{4})\s*$', str(complemento))
    return m.group(1) if m else None

def parse_valor(v):
    if not v: return 0.0
    v = str(v).strip().replace('.', '').replace(',', '.')
    try: return round(abs(float(v)), 2)
    except: return 0.0

def parse_date(d_str):
    try: return datetime.strptime(str(d_str).strip(), '%d/%m/%Y')
    except: return None

def parse_filename(fname):
    dt, amt = None, None
    d_match = re.search(r'(\d{2}[-_\/]?\d{2}[-_\/]?\d{4}|\d{8})', fname)
    if d_match:
        d_str = d_match.group(1).replace('_','').replace('-','').replace('/','')
        try: dt = datetime.strptime(d_str, '%d%m%Y')
        except: pass
    a_match = re.search(r'(\d{1,3}[.,]\d{2})\b', fname)
    if a_match:
        try: amt = float(a_match.group(1).replace('.','').replace(',','.'))
        except: pass
    return dt, amt

def build_comprovante_lookup(zip_path, tolerance_days):
    lookup = {}
    with zipfile.ZipFile(zip_path, 'r') as zf:
        for info in zf.infolist():
            fname = os.path.basename(info.filename)
            if not fname.lower().endswith('.pdf'): continue
            dt, amt = parse_filename(fname)
            if dt and amt:
                for delta in range(-tolerance_days, tolerance_days + 1):
                    key = (dt.date() + timedelta(days=delta), round(abs(amt), 2))
                    lookup.setdefault(key, []).append({'filename': info.filename, 'delta': delta})
    return lookup

def lookup_comprovante(date_obj, valor, lookup):
    if not date_obj: return None
    key = (date_obj.date(), round(valor, 2))
    results = lookup.get(key, [])
    if not results: return None
    results.sort(key=lambda x: abs(x['delta']))
    return results[0]

def extract_cnpj_from_pdf(filename, zip_path):
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf: data = zf.read(filename)
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            text = '\n'.join(p.extract_text() or '' for p in pdf.pages[:2])
        return extract_cnpj(text)
    except: return None

def batch_lookup_cnpjs(unknown_cnpjs, task_id, cnpj_cache):
    cfg = load_config()
    if not cfg.get('receita_ws_enabled', True):
        return cnpj_cache
    to_fetch = [c for c in unknown_cnpjs if c not in cnpj_cache and len(c) == 14]
    if not to_fetch:
        return cnpj_cache
    rpm   = cfg.get('receita_ws_rpm', 3)
    delay = 60.0 / rpm
    total = len(to_fetch)
    found = 0
    tasks[task_id]['log'].append(f"Consultando ReceitaWS: {total} CNPJs desconhecidos...")
    with get_db() as conn:
        for i, cnpj in enumerate(to_fetch):
            try:
                r = requests.get(f"https://receitaws.com.br/v1/cnpj/{cnpj}", timeout=6)
                if r.status_code == 200:
                    data = r.json()
                    if data.get('status') != 'ERROR':
                        nome = data.get('nome', '').strip()
                        cnpj_cache[cnpj] = nome
                        conn.execute('INSERT OR REPLACE INTO cnpj_cache (cnpj, nome) VALUES (?,?)', (cnpj, nome))
                        found += 1
                    else:
                        cnpj_cache[cnpj] = None
                        conn.execute('INSERT OR REPLACE INTO cnpj_cache (cnpj, nome) VALUES (?,NULL)', (cnpj,))
                elif r.status_code == 429:
                    tasks[task_id]['log'].append(f"ReceitaWS: limite atingido em {i+1}/{total}. Aguardando...")
                    time.sleep(20)
            except:
                cnpj_cache[cnpj] = None
            if (i + 1) % 3 == 0 or i == total - 1:
                tasks[task_id]['log'].append(f"ReceitaWS: {i+1}/{total} consultados, {found} encontrados.")
            time.sleep(delay)
    return cnpj_cache

def build_excel(results, new_fornecedores, out_path, cfg):
    wb = Workbook()
    ws = wb.active
    ws.title = "RESULTADO"
    cols = [
        ('Data', 12), ('Historico', 20), ('Complemento', 45), ('Valor', 12),
        ('Debito', 15), ('Credito', 15), ('CODIGO', 10), ('PARTICIPANTE', 35),
        ('CNPJ_IDENTIFICADO', 20), ('FOLHA', 8), ('CHEQUE_NUM', 12),
        ('COMPROVANTE', 35), ('STATUS', 14), ('ORIGEM', 14), ('CONFIRMAR', 14),
    ]
    col_keys = ['Data', 'Historico', 'Complemento', 'Valor', 'Debito', 'Credito',
                'CODIGO', 'PARTICIPANTE', 'CNPJ_IDENTIFICADO', 'FOLHA', 'CHEQUE_NUM',
                'COMPROVANTE', 'STATUS', 'ORIGEM', 'CONFIRMAR']
    header_fill   = PatternFill(start_color="0D0F12", end_color="0D0F12", fill_type="solid")
    header_font   = Font(bold=True, color="00D4FF", size=10)
    header_border = Border(bottom=Side(style='thin', color="00D4FF"))
    for i, (col, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = header_border
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    code_opts = '"' + cfg['codigo_fornecedor'] + ',' + cfg['codigo_folha'] + ',' + cfg['codigo_nao_identificado'] + '"'
    for dv_def, col_letter in [
        (DataValidation(type="list", formula1=code_opts, allow_blank=True, showDropDown=False), 'G'),
        (DataValidation(type="list", formula1='"OK,REVISAR,NAO ENCONTRADO"', allow_blank=True, showDropDown=False), 'M'),
        (DataValidation(type="list", formula1='"SIM,REVISAR,NAO"', allow_blank=True, showDropDown=False), 'O'),
    ]:
        dv_def.sqref = f"{col_letter}2:{col_letter}{len(results)+1}"
        ws.add_data_validation(dv_def)
    green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    blue_fill   = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    grey_fill   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for r in results:
        ws.append([r.get(k, '') for k in col_keys])
        if r['STATUS'] == 'TRANSFERENCIA_PROPRIA': fill = grey_fill
        elif r['STATUS'] == 'OK':                  fill = green_fill
        elif r['STATUS'] == 'REVISAR':             fill = blue_fill if r.get('ORIGEM') == 'RECEITA_WS' else yellow_fill
        else:                                       fill = red_fill
        for cell in ws[ws.max_row]:
            cell.fill      = fill
            cell.alignment = Alignment(vertical='center')
    if new_fornecedores:
        ws2 = wb.create_sheet("NOVOS FORNECEDORES")
        ws2.append(['CNPJ', 'NOME (ReceitaWS)', 'ADICIONAR AO CADASTRO?'])
        for cell in ws2[1]:
            cell.font  = Font(bold=True, color="00D4FF")
            cell.fill  = PatternFill(start_color="0D0F12", end_color="0D0F12", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        dv2 = DataValidation(type="list", formula1='"SIM,NAO"', allow_blank=False, showDropDown=False)
        dv2.sqref = f"C2:C{len(new_fornecedores)+1}"
        ws2.add_data_validation(dv2)
        for cnpj, nome in new_fornecedores.items():
            ws2.append([cnpj, nome, 'SIM'])
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 55
        ws2.column_dimensions['C'].width = 22
    wb.save(out_path)

def run_fam_etl(task_id, sci_path, zip_path):
    tasks[task_id]['status'] = 'RUNNING'
    cfg = load_config()
    logger.info(f"Run started: task={task_id}")

    def log(msg):
        tasks[task_id]['log'].append(msg)
        logger.info(msg)

    try:
        forn_map, func_names, cheques_map, cnpj_cache = build_maps()
        log(f"Dados carregados: {len(forn_map)} fornecedores, {len(func_names)} funcionarios, {len(cheques_map)} cheques, {len(cnpj_cache)} CNPJs em cache.")

        log("Lendo SCI Bank CSV...")
        try: df = pd.read_csv(sci_path, sep=';', encoding='utf-8', dtype=str)
        except: df = pd.read_csv(sci_path, sep=';', encoding='latin-1', dtype=str)
        std_cols = ['Transacao','Chave','Lote','Data','Debito','Participante a debito',
                    'Credito','Participante a credito','Valor','Historico','Complemento','No Doc.']
        if len(df.columns) == len(std_cols) + 1:
            df = df.iloc[:, 1:]
        df.columns = std_cols[:len(df.columns)]
        df.fillna('', inplace=True)
        df['Valor_Float'] = df['Valor'].apply(parse_valor)
        log(f"{len(df)} transacoes carregadas.")

        log("Indexando Comprovantes ZIP...")
        tolerance = cfg.get('comprovante_tolerance_days', 1)
        comp_lookup = build_comprovante_lookup(zip_path, tolerance)
        log(f"{len(comp_lookup)} chaves de comprovante indexadas.")

        log("Identificando CNPJs desconhecidos...")
        unknown_cnpjs = set()
        for _, row in df.iterrows():
            cnpj = extract_cnpj(str(row.get('Complemento', '')))
            if cnpj and len(cnpj) == 14 and cnpj not in forn_map and cnpj != FAM_OWN_CNPJ:
                unknown_cnpjs.add(cnpj)
        log(f"{len(unknown_cnpjs)} CNPJs desconhecidos encontrados.")

        if cfg.get('receita_ws_enabled', True) and unknown_cnpjs:
            cnpj_cache = batch_lookup_cnpjs(unknown_cnpjs, task_id, cnpj_cache)

        log("Enriquecendo transacoes...")
        results = []
        new_fornecedores = {}
        CODE_FORN  = cfg['codigo_fornecedor']
        CODE_FOLHA = cfg['codigo_folha']
        CODE_NAO   = cfg['codigo_nao_identificado']

        stat_counts = {'self': 0, 'cadastro': 0, 'receita': 0, 'comp_name': 0, 'folha': 0, 'cheque': 0, 'pdf': 0, 'nao': 0}

        for idx, row in df.iterrows():
            comp_text = str(row.get('Complemento', ''))
            hist_text = str(row.get('Historico', ''))
            res = {
                'Data': row['Data'], 'Historico': hist_text, 'Complemento': comp_text,
                'Valor': row['Valor'], 'Debito': row['Debito'], 'Credito': row['Credito'],
                'CODIGO': CODE_NAO, 'PARTICIPANTE': '', 'CNPJ_IDENTIFICADO': '',
                'FOLHA': '', 'CHEQUE_NUM': '', 'COMPROVANTE': '',
                'STATUS': 'NAO ENCONTRADO', 'ORIGEM': '', 'CONFIRMAR': ''
            }

            # ── Self-transfer detection ──────────────────────────────────────
            if is_self_transfer(comp_text):
                res['STATUS']      = 'TRANSFERENCIA_PROPRIA'
                res['PARTICIPANTE'] = 'FAM METAL (TRANSFERENCIA PROPRIA)'
                res['ORIGEM']       = 'SELF'
                res['CODIGO']       = ''
                stat_counts['self'] += 1
                results.append(res)
                continue

            cnpj = extract_cnpj(comp_text)
            cpf  = extract_cpf(comp_text)

            # ── A: CNPJ lookup ───────────────────────────────────────────────
            if cnpj and cnpj in forn_map:
                res['PARTICIPANTE']      = forn_map[cnpj]
                res['CNPJ_IDENTIFICADO'] = cnpj
                res['ORIGEM']            = 'CADASTRO'
                res['CODIGO']            = CODE_FORN
                stat_counts['cadastro'] += 1

            elif cnpj and cnpj in cnpj_cache and cnpj_cache[cnpj]:
                res['PARTICIPANTE']      = cnpj_cache[cnpj]
                res['CNPJ_IDENTIFICADO'] = cnpj
                res['ORIGEM']            = 'RECEITA_WS'
                res['CODIGO']            = CODE_FORN
                new_fornecedores[cnpj]   = cnpj_cache[cnpj]
                stat_counts['receita'] += 1

            elif cnpj:
                res['CNPJ_IDENTIFICADO'] = cnpj

            elif cpf:
                res['CNPJ_IDENTIFICADO'] = 'CPF:' + cpf

            # ── B: Name extraction from Complemento text ─────────────────────
            # Runs for ALL rows — fills PARTICIPANTE when CNPJ lookup missed
            if not res['PARTICIPANTE']:
                fallback_name, origin_hint = extract_name_from_complemento(comp_text)
                if fallback_name:
                    res['PARTICIPANTE'] = '(?) ' + fallback_name
                    res['ORIGEM']       = origin_hint or 'COMPLEMENTO'
                    res['CODIGO']       = CODE_FORN
                    stat_counts['comp_name'] += 1

            # ── C: PIX folha detection (employee name match) ─────────────────
            pix_name = extract_pix_name(comp_text)
            if pix_name and func_names:
                is_folha, reason = match_employee_name(pix_name, func_names)
                if is_folha:
                    res['FOLHA']  = 'SIM'
                    res['CODIGO'] = CODE_FOLHA
                    res['ORIGEM'] = 'FOLHA'
                    stat_counts['folha'] += 1
                    if not res['PARTICIPANTE'] or res['PARTICIPANTE'].startswith('(?)'):
                        res['PARTICIPANTE'] = reason

            # ── D: Cheque lookup ─────────────────────────────────────────────
            if 'CHEQUE' in comp_text.upper():
                ch_num = extract_cheque_number(comp_text)
                if ch_num and ch_num in cheques_map:
                    portador = cheques_map[ch_num]
                    res['CHEQUE_NUM']   = ch_num
                    res['PARTICIPANTE'] = portador
                    res['ORIGEM']       = 'CHEQUE'
                    stat_counts['cheque'] += 1
                    if any(k in portador.upper() for k in ['SALARIO','FAM METAL']):
                        res['CODIGO'] = CODE_FOLHA
                        res['FOLHA']  = 'SIM'
                    else:
                        res['CODIGO'] = CODE_FORN

            # ── E: Comprovante match by date+value ───────────────────────────
            dt_obj = parse_date(row['Data'])
            comp_match = lookup_comprovante(dt_obj, row['Valor_Float'], comp_lookup)
            if comp_match:
                res['COMPROVANTE'] = comp_match['filename']
                if not res['PARTICIPANTE'] or res['PARTICIPANTE'].startswith('(?)'):
                    pdf_cnpj = extract_cnpj_from_pdf(comp_match['filename'], zip_path)
                    if pdf_cnpj and pdf_cnpj in forn_map:
                        res['PARTICIPANTE']      = forn_map[pdf_cnpj]
                        res['CNPJ_IDENTIFICADO'] = pdf_cnpj
                        res['ORIGEM']            = 'PDF'
                        res['CODIGO']            = CODE_FORN
                        stat_counts['pdf'] += 1

            # ── Status ───────────────────────────────────────────────────────
            has_participant = bool(res['PARTICIPANTE']) and res['PARTICIPANTE'] != '(?) '
            has_comprovante = bool(res['COMPROVANTE'])

            # Strip the (?) prefix from PARTICIPANTE if we also have a comprovante
            # — confirmed enough to be OK
            if has_participant and res['PARTICIPANTE'].startswith('(?)') and has_comprovante:
                res['PARTICIPANTE'] = res['PARTICIPANTE'][4:]

            if has_participant and has_comprovante:
                res['STATUS'] = 'OK'
            elif has_participant or has_comprovante:
                res['STATUS'] = 'REVISAR'
            else:
                stat_counts['nao'] += 1

            results.append(res)

        if new_fornecedores:
            with get_db() as conn:
                for cnpj, nome in new_fornecedores.items():
                    conn.execute('INSERT OR IGNORE INTO fornecedores (cnpj, nome, source, confirmed) VALUES (?,?,?,0)',
                                (cnpj, nome, 'RECEITA_WS'))

        log("Gerando Excel...")
        out_path = os.path.join(app.config['OUTPUT_FOLDER'], f"{task_id}.xlsx")
        build_excel(results, new_fornecedores, out_path, cfg)

        ok_c   = sum(1 for r in results if r['STATUS'] == 'OK')
        rev_c  = sum(1 for r in results if r['STATUS'] == 'REVISAR')
        nao_c  = sum(1 for r in results if r['STATUS'] == 'NAO ENCONTRADO')
        self_c = sum(1 for r in results if r['STATUS'] == 'TRANSFERENCIA_PROPRIA')
        total  = len(results)

        with get_db() as conn:
            conn.execute('INSERT INTO run_history (total_rows,ok_count,revisar_count,nao_encontrado_count,novos_fornecedores) VALUES (?,?,?,?,?)',
                        (total, ok_c, rev_c, nao_c, len(new_fornecedores)))

        tasks[task_id]['status'] = 'DONE'
        tasks[task_id]['file']   = out_path
        log(f"--- RESUMO ---")
        log(f"OK: {ok_c} ({ok_c/total*100:.1f}%) | REVISAR: {rev_c} ({rev_c/total*100:.1f}%) | NAO ENCONTRADO: {nao_c} ({nao_c/total*100:.1f}%) | PROPRIA: {self_c}")
        log(f"Origens — CADASTRO: {stat_counts['cadastro']} | RECEITA_WS: {stat_counts['receita']} | COMPLEMENTO: {stat_counts['comp_name']} | FOLHA: {stat_counts['folha']} | CHEQUE: {stat_counts['cheque']} | PDF: {stat_counts['pdf']} | SELF: {stat_counts['self']} | NAO: {stat_counts['nao']}")
        log(f"Novos fornecedores via API: {len(new_fornecedores)}")
        log("Processamento concluido com sucesso!")

    except Exception as e:
        import traceback
        tasks[task_id]['status'] = 'ERROR'
        tasks[task_id]['log'].append(f"ERRO: {str(e)}")
        tasks[task_id]['log'].append(traceback.format_exc())
        logger.error(f"Run failed: {e}")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload/sci', methods=['POST'])
def upload_sci():
    if 'file' not in request.files: return jsonify({"error": "No file"}), 400
    f = request.files['file']
    path = os.path.join(app.config['UPLOAD_FOLDER'], f"sci_{uuid.uuid4().hex}.csv")
    f.save(path)
    return jsonify({"path": path})

@app.route('/upload/zip', methods=['POST'])
def upload_zip():
    if 'file' not in request.files: return jsonify({"error": "No file"}), 400
    f = request.files['file']
    path = os.path.join(app.config['UPLOAD_FOLDER'], f"comp_{uuid.uuid4().hex}.zip")
    f.save(path)
    return jsonify({"path": path})

@app.route('/processar', methods=['POST'])
def processar():
    data     = request.json
    sci_path = data.get('sci_path')
    zip_path = data.get('zip_path')
    if not sci_path or not zip_path: return jsonify({"error": "Missing files"}), 400
    task_id = uuid.uuid4().hex
    tasks[task_id] = {"status": "PENDING", "log": [], "file": None}
    threading.Thread(target=run_fam_etl, args=(task_id, sci_path, zip_path)).start()
    return jsonify({"task_id": task_id})

@app.route('/status/<task_id>')
def status(task_id):
    if task_id not in tasks: return jsonify({"error": "Invalid task"}), 404
    return jsonify(tasks[task_id])

@app.route('/download/<task_id>')
def download(task_id):
    if task_id not in tasks or tasks[task_id]['status'] != 'DONE':
        return jsonify({"error": "Not ready"}), 400
    return send_file(tasks[task_id]['file'], as_attachment=True, download_name='FAM_Resultado.xlsx')

@app.route('/data/<ref_type>', methods=['GET'])
def get_data(ref_type):
    if ref_type not in ['funcionarios', 'fornecedores', 'cheques']:
        return jsonify({"error": "Invalid"}), 400
    with get_db() as conn:
        if ref_type == 'fornecedores':
            rows = conn.execute('SELECT cnpj as CNPJ, nome as NOME, source, confirmed FROM fornecedores').fetchall()
        elif ref_type == 'funcionarios':
            rows = conn.execute('SELECT nome as NOME, confirmed FROM funcionarios').fetchall()
        else:
            rows = conn.execute('SELECT numero as Numero, portador as Portador FROM cheques').fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/data/<ref_type>', methods=['POST'])
def update_data(ref_type):
    if ref_type not in ['funcionarios', 'fornecedores', 'cheques']:
        return jsonify({"error": "Invalid"}), 400
    if 'file' not in request.files: return jsonify({"error": "No file"}), 400
    f = request.files['file']
    path = os.path.join(app.config['DATA_FOLDER'], f"{ref_type}_upload.csv")
    f.save(path)
    df = pd.read_csv(path, dtype=str).fillna('')
    with get_db() as conn:
        if ref_type == 'fornecedores':
            for _, row in df.iterrows():
                cnpj = re.sub(r'[^\d]', '', row.get('CNPJ', ''))
                nome = row.get('NOME', '').strip()
                if cnpj and nome:
                    conn.execute('INSERT OR REPLACE INTO fornecedores (cnpj, nome, source, confirmed) VALUES (?,?,?,1)', (cnpj, nome, 'CADASTRO'))
        elif ref_type == 'funcionarios':
            for _, row in df.iterrows():
                nome = row.get('NOME', '').strip()
                if nome:
                    conn.execute('INSERT OR IGNORE INTO funcionarios (nome) VALUES (?)', (nome,))
        else:
            for _, row in df.iterrows():
                num = row.get('Numero', '').strip()
                por = row.get('Portador', '').strip()
                if num and por:
                    conn.execute('INSERT OR REPLACE INTO cheques (numero, portador) VALUES (?,?)', (num, por))
    return jsonify({"success": True})

@app.route('/config', methods=['GET'])
def get_config():
    return jsonify(load_config())

@app.route('/config', methods=['POST'])
def update_config():
    data = request.json
    cfg  = load_config()
    for k in ['codigo_fornecedor','codigo_folha','codigo_nao_identificado',
              'comprovante_tolerance_days','receita_ws_enabled','receita_ws_rpm']:
        if k in data:
            cfg[k] = data[k]
    save_config(cfg)
    return jsonify({"success": True, "config": cfg})

@app.route('/history', methods=['GET'])
def get_history():
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM run_history ORDER BY run_date DESC LIMIT 20').fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/confirm_fornecedores', methods=['POST'])
def confirm_fornecedores():
    data  = request.json
    cnpjs = data.get('cnpjs', [])
    with get_db() as conn:
        for cnpj in cnpjs:
            conn.execute('UPDATE fornecedores SET confirmed=1 WHERE cnpj=?', (cnpj,))
    return jsonify({"success": True, "confirmed": len(cnpjs)})

if __name__ == '__main__':
    init_db()
    migrate_csvs_to_db()
    cfg = load_config()
    logger.info(f"FAM App starting on port {cfg['port']}")
    app.run(host='0.0.0.0', port=cfg['port'], debug=False)
