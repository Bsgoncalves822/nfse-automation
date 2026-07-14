import os
import glob
import json
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NS = {'nfse': 'http://www.sped.fazenda.gov.br/nfse'}

CIDADES_LOOKUP_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config', 'cidades_lookup.json')

def load_cidades():
    try:
        with open(CIDADES_LOOKUP_PATH, encoding='utf-8') as f:
            return json.load(f)
    except:
        return {}

CIDADES = load_cidades()

def get_text(root, path):
    el = root.find(path, NS)
    return el.text.strip() if el is not None and el.text else ''

def get_float(root, path):
    val = get_text(root, path)
    try:
        return float(val.replace(',', '.'))
    except:
        return 0.0

def parse_xml(xml_path):
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()

        n_nfse    = get_text(root, './/nfse:nNFSe')
        x_loc_emi = get_text(root, './/nfse:xLocEmi')
        dh_emi    = get_text(root, './/nfse:dhEmi')
        serie     = get_text(root, './/nfse:serie')

        cnpj_emit = get_text(root, './/nfse:emit/nfse:CNPJ')
        nome_emit = get_text(root, './/nfse:emit/nfse:xNome')
        uf        = get_text(root, './/nfse:emit/nfse:enderNac/nfse:UF') or 'SC'
        cmun      = get_text(root, './/nfse:emit/nfse:enderNac/nfse:cMun') or ''

        cnpj_clean = ''.join(filter(str.isdigit, cnpj_emit))
        if len(cnpj_clean) == 14:
            cnpj_fmt = f'{cnpj_clean[:2]}.{cnpj_clean[2:5]}.{cnpj_clean[5:8]}/{cnpj_clean[8:12]}-{cnpj_clean[12:]}'
        elif len(cnpj_clean) == 11:
            cnpj_fmt = f'{cnpj_clean[:3]}.{cnpj_clean[3:6]}.{cnpj_clean[6:9]}-{cnpj_clean[9:]}'
        else:
            cnpj_fmt = cnpj_emit

        try:
            dt       = datetime.fromisoformat(dh_emi[:19])
            data_fmt = dt.strftime('%d/%m/%Y')
            data_txt = dt.strftime('%Y%m%d')
        except:
            data_fmt = dh_emi[:10]
            data_txt = dh_emi[:10].replace('-', '')

        v_serv = get_float(root, './/nfse:vServ')
        if v_serv == 0:
            v_serv = get_float(root, './/nfse:vBC')

        x_trib_nac = get_text(root, './/nfse:xTribNac')
        natureza   = x_trib_nac[:90] if x_trib_nac else ''

        v_ret_irrf = get_float(root, './/nfse:vRetIRRF')
        v_ret_csll = get_float(root, './/nfse:vRetCSLL')
        v_pis      = get_float(root, './/nfse:vPis')
        v_cofins   = get_float(root, './/nfse:vCofins')
        v_ret_inss = get_float(root, './/nfse:vRetINSS')
        v_ret_cp   = get_float(root, './/nfse:vRetCP')
        v_issqn    = get_float(root, './/nfse:vISSQN')

        if v_ret_cp > 0:
            inss_val  = v_ret_cp
            inss_aliq = 3.5
        elif v_ret_inss > 0:
            inss_val  = v_ret_inss
            inss_aliq = 11.0
        else:
            inss_val  = 0.0
            inss_aliq = 0.0

        total_ret   = v_ret_irrf + v_ret_csll + v_pis + v_cofins + inss_val
        conta       = '1933000' if uf == 'SC' else '2933001'
        cidade_code = CIDADES.get(cmun, cmun)

        return {
            'data':        data_fmt,
            'data_txt':    data_txt,
            'numero':      n_nfse,
            'serie':       serie,
            'cnpj':        cnpj_fmt,
            'cnpj_digits': cnpj_clean,
            'nome':        nome_emit,
            'cidade':      x_loc_emi,
            'cmun':        cmun,
            'cidade_code': cidade_code,
            'uf':          uf,
            'conta':       conta,
            'natureza':    natureza,
            'v_serv':      v_serv,
            'v_iss':       v_issqn,
            'v_pis':       v_pis,
            'v_cofins':    v_cofins,
            'v_ir':        v_ret_irrf,
            'v_csll':      v_ret_csll,
            'v_inss':      inss_val,
            'inss_aliq':   inss_aliq,
            'total_ret':   total_ret,
        }
    except Exception as e:
        print(f'[AVISO] Erro ao parsear {xml_path}: {e}')
        return None

def generate_fiscal(company_name, company_dir, month):
    federal_xml_dir = os.path.join(company_dir, 'federal', 'xmls')
    if not os.path.exists(federal_xml_dir):
        print(f'[INFO] Sem XMLs federais para {company_name}')
        return None

    xml_files = glob.glob(os.path.join(federal_xml_dir, '*.xml'))
    if not xml_files:
        print(f'[INFO] Pasta vazia: {federal_xml_dir}')
        return None

    rows = []
    for xml_path in sorted(xml_files):
        data = parse_xml(xml_path)
        if data:
            rows.append(data)

    if not rows:
        return None

    rows.sort(key=lambda r: (r['data'], r['numero']))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Retencoes'

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', start_color='C0392B')
    title_font  = Font(name='Arial', bold=True, size=11, color='C0392B')
    total_font  = Font(name='Arial', bold=True, size=10)
    total_fill  = PatternFill('solid', start_color='F9E79F')
    normal_font = Font(name='Arial', size=10)
    center      = Alignment(horizontal='center', vertical='center')
    left        = Alignment(horizontal='left',   vertical='center')
    right       = Alignment(horizontal='right',  vertical='center')
    thin        = Side(style='thin', color='BFBFBF')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:N1')
    ws['A1']           = f'NOTAS FISCAIS COM RETENCAO FEDERAL — {company_name}'
    ws['A1'].font      = title_font
    ws['A1'].alignment = left
    ws.row_dimensions[1].height = 20

    ws.merge_cells('A2:N2')
    ws['A2']           = f'Periodo: {month}   |   Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font      = Font(name='Arial', size=9, color='7F8C8D')
    ws['A2'].alignment = left
    ws.row_dimensions[2].height = 14

    headers = [
        'Nº NFSe', 'Emissão', 'CNPJ/CPF Emitente', 'Razão Emitente',
        'Vl. Serviço', 'ISS Ret.', 'Pis Ret.', 'Cofins Ret.',
        'IR Ret.', 'CSLL Ret.', 'INSS Ret.', 'Aliq. INSS %', 'Total Retido', 'Natureza'
    ]
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=3, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[3].height = 16

    money_fmt = '#,##0.00'
    for i, r in enumerate(rows, start=4):
        fill = PatternFill('solid', start_color='FDFEFE') if i % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
        values = [
            r['numero'], r['data'], r['cnpj'], r['nome'],
            r['v_serv'], r['v_iss'], r['v_pis'], r['v_cofins'],
            r['v_ir'], r['v_csll'], r['v_inss'],
            r['inss_aliq'] if r['inss_aliq'] > 0 else None,
            r['total_ret'], r['natureza']
        ]
        for col, val in enumerate(values, 1):
            cell        = ws.cell(row=i, column=col, value=val)
            cell.font   = normal_font
            cell.fill   = fill
            cell.border = border
            if col in [5, 6, 7, 8, 9, 10, 11, 13]:
                cell.number_format = money_fmt
                cell.alignment     = right
            elif col == 12:
                cell.number_format = '0.00'
                cell.alignment     = center
            elif col in [1, 2]:
                cell.alignment = center
            else:
                cell.alignment = left

    total_row = len(rows) + 4
    ws.cell(row=total_row, column=1, value='TOTAIS').font      = total_font
    ws.cell(row=total_row, column=1).fill                      = total_fill
    ws.cell(row=total_row, column=1).alignment                 = left
    ws.cell(row=total_row, column=1).border                    = border

    total_cols = {5: 'v_serv', 6: 'v_iss', 7: 'v_pis', 8: 'v_cofins',
                  9: 'v_ir', 10: 'v_csll', 11: 'v_inss', 13: 'total_ret'}
    for col in range(2, 15):
        cell        = ws.cell(row=total_row, column=col)
        cell.fill   = total_fill
        cell.border = border
        if col in total_cols:
            cell.value         = sum(r[total_cols[col]] for r in rows)
            cell.number_format = money_fmt
            cell.alignment     = right
            cell.font          = total_font

    col_letters = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']
    col_widths  = [10, 12, 22, 40, 14, 12, 12, 14, 12, 12, 12, 12, 14, 60]
    for letter, w in zip(col_letters, col_widths):
        ws.column_dimensions[letter].width = w

    ws.freeze_panes = 'A4'

    fiscal_dir = os.path.join(company_dir, 'fiscal')
    os.makedirs(fiscal_dir, exist_ok=True)
    safe_name = company_name.replace('/', '_').replace('\\', '_').replace(':', '_')
    out_path  = os.path.join(fiscal_dir, f'fiscal_{safe_name}_{month}.xlsx')
    wb.save(out_path)
    print(f'[OK] Fiscal XLSX salvo: {out_path}')
    return out_path

def generate_fiscal_txt(company_name, company_dir, month):
    federal_xml_dir = os.path.join(company_dir, 'federal', 'xmls')
    if not os.path.exists(federal_xml_dir):
        return None

    xml_files = glob.glob(os.path.join(federal_xml_dir, '*.xml'))
    if not xml_files:
        return None

    rows = []
    for xml_path in sorted(xml_files):
        data = parse_xml(xml_path)
        if data:
            rows.append(data)

    if not rows:
        return None

    rows.sort(key=lambda r: (r['data'], r['numero']))

    def fmt(v):
        # Always returns a value, '0' when zero
        if v == 0:
            return '0'
        s = f'{v:.2f}'
        if '.' in s:
            s = s.rstrip('0').rstrip('.')
        return s

    def fmt_blank(v):
        # Returns empty string when zero
        if v == 0:
            return ''
        s = f'{v:.2f}'
        if '.' in s:
            s = s.rstrip('0').rstrip('.')
        return s

    lines = []
    for i, r in enumerate(rows, start=1):
        inss_cprb  = '1' if r['inss_aliq'] == 3.5 else '0'
        has_ir     = r['v_ir'] > 0
        has_inss   = r['v_inss'] > 0
        ir_aliq    = '1.5' if has_ir else '0'
        reinf_code = '"100000003 "' if not has_ir else '"100000003"'

        # Natureza rendimento codes
        # IR = 15018 area, PIS/COFINS/CSLL = 470, INSS = 470
        nat_ir     = '15018' if has_ir else ''

        fields = [''] * 242  # 242 fields to match her format

        fields[0]   = str(i)
        fields[1]   = r['cnpj']
        fields[2]   = r['cidade_code']
        fields[3]   = r['uf']
        fields[4]   = r['data_txt']
        fields[5]   = r['numero']
        fields[6]   = r['numero']
        fields[7]   = 'NFSE'
        fields[8]   = '1'
        fields[9]   = fmt_blank(r['v_serv'])
        fields[10]  = '0'
        fields[11]  = '0'
        fields[12]  = '0'
        # field 13 = empty (observacao)
        fields[14]  = fmt(r['v_ir'])        # base IR
        fields[15]  = ir_aliq               # aliq IR
        fields[16]  = fmt(r['v_ir'])        # valor IR
        fields[17]  = 'V'
        # fields 18-19 empty
        fields[20]  = '0'
        # fields 21-40 empty
        fields[41]  = fmt_blank(r['v_inss'])  # INSS retido
        # fields 42-43 empty
        fields[44]  = 'R'
        # fields 45-97 empty
        fields[98]  = r['conta']
        fields[99]  = fmt_blank(r['v_pis'])    # PIS
        fields[100] = fmt_blank(r['v_cofins']) # COFINS
        fields[101] = fmt_blank(r['v_csll'])   # CSLL
        # fields 102-105 empty
        fields[106] = fmt_blank(r['v_inss']) if has_inss else ''  # INSS base
        # fields 107-147 empty
        fields[148] = '00'
        # fields 149-168 empty
        fields[169] = '1708' if has_ir else ''
        # fields 170-221 empty
        fields[222] = '"1"'
        fields[223] = '"900189501274"'
        fields[224] = reinf_code
        fields[225] = f'"{inss_cprb}"'
        # fields 226-233 empty
        fields[234] = fmt_blank(r['v_inss']) if has_inss else ''  # INSS valor
        fields[235] = fmt_blank(r['v_serv']) if not has_inss else ''  # base serv when no INSS
        # fields 236 empty
        fields[237] = '470'   # nat PIS
        fields[238] = '470'   # nat COFINS
        fields[239] = '470'   # nat CSLL
        # fields 240-241 empty

        lines.append(','.join(fields))

    fiscal_dir = os.path.join(company_dir, 'fiscal')
    os.makedirs(fiscal_dir, exist_ok=True)
    safe_name = company_name.replace('/', '_').replace('\\', '_').replace(':', '_')
    out_path  = os.path.join(fiscal_dir, f'fiscal_{safe_name}_{month}.txt')
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    print(f'[OK] Fiscal TXT salvo: {out_path}')
    return out_path

def _name_matches(folder_name, filter_names):
    """Match folder name against filter list — handles name drift and CNPJ suffix on disk."""
    folder_upper = folder_name.upper()
    for n in filter_names:
        n_upper = n.upper()
        # exact match, or folder starts with the filter name, or filter name is contained in folder
        if folder_upper == n_upper or folder_upper.startswith(n_upper) or n_upper in folder_upper:
            return True
    return False

def generate_fiscal_all(filter_names=None):
    settings_path = Path(__file__).parent / 'config' / 'settings.json'
    with open(settings_path, encoding='utf-8') as f:
        settings = json.load(f)
    base = Path(settings['downloads_path'])

    generated = []
    for accountant_dir in base.iterdir():
        if not accountant_dir.is_dir():
            continue
        for company_dir in accountant_dir.iterdir():
            if not company_dir.is_dir():
                continue
            if filter_names and not _name_matches(company_dir.name, filter_names):
                continue
            for month_dir in company_dir.iterdir():
                if not month_dir.is_dir():
                    continue
                result = generate_fiscal(company_dir.name, str(month_dir), month_dir.name)
                if result:
                    generated.append(result)
                generate_fiscal_txt(company_dir.name, str(month_dir), month_dir.name)
    return generated

if __name__ == '__main__':
    generate_fiscal_all()