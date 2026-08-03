"""
NFSe RetenÃ§Ã£o Parser â€” ORPROCON
-------------------------------
Parses all NFSe XMLs from Questor export folders, groups by tomador (client company),
copies matching DANFSe PDFs, and generates a consolidated Excel retention report.

Usage:
    python nfse_retencao.py --xml-dir "C:\\path\\to\\XMLs\\Padrao" \
                            --pdf-dir "C:\\path\\to\\DANFSes" \
                            --out-dir "C:\\path\\to\\output"

All three source folders can be passed if needed:
    --xml-dir can be specified multiple times, or point to a parent folder
    (the script walks subdirectories automatically).
"""

import argparse
import shutil
import sys
from pathlib import Path
from collections import defaultdict
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NS = "http://www.sped.fazenda.gov.br/nfse"


# â”€â”€ XML helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def tag(name):
    return f"{{{NS}}}{name}"


def findtext(el, *path, default=""):
    for p in path:
        if el is None:
            return default
        el = el.find(tag(p))
    if el is None or el.text is None:
        return default
    return el.text.strip()


def money(text):
    try:
        return float(text.replace(",", ".")) if text else 0.0
    except (ValueError, AttributeError):
        return 0.0


# â”€â”€ Parse a single XML â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def parse_xml(path: Path) -> dict | None:
    try:
        tree = ET.parse(path)
        r = tree.getroot()
        # handle both bare NFSe and wrapped CompNFSe
        inf = r.find(tag("infNFSe"))
        if inf is None:
            inf = r.find(f".//{tag('infNFSe')}")
        if inf is None:
            return None

        dps_inf = inf.find(f".//{tag('infDPS')}")
        if dps_inf is None:
            return None

        emit = inf.find(tag("emit"))
        toma = dps_inf.find(tag("toma"))
        inf_vals = inf.find(tag("valores"))
        dps_vals = dps_inf.find(tag("valores"))
        trib = dps_vals.find(tag("trib")) if dps_vals else None
        trib_mun = trib.find(tag("tribMun")) if trib else None
        trib_fed = trib.find(tag("tribFed")) if trib else None
        ret_fed = trib_fed.find(tag("retencoes")) if trib_fed else None

        # â”€â”€ tomador (client) â”€â”€
        toma_cnpj = findtext(toma, "CNPJ") or findtext(toma, "CPF")
        toma_nome = findtext(toma, "xNome")

        # â”€â”€ prestador (service provider) â”€â”€
        prest_cnpj = findtext(emit, "CNPJ") or findtext(emit, "CPF")
        prest_nome = findtext(emit, "xNome")

        # â”€â”€ nota metadata â”€â”€
        n_nfse   = findtext(inf, "nNFSe")
        dh_proc  = findtext(inf, "dhProc")
        d_compet = findtext(dps_inf, "dCompet")
        xdesc    = findtext(dps_inf, "serv", "cServ", "xDescServ")
        xloc     = findtext(inf, "xLocEmi")

        # â”€â”€ valores â”€â”€
        v_serv     = money(findtext(dps_vals, "vServPrest", "vServ"))
        v_bc       = money(findtext(inf_vals, "vBC"))
        v_issqn    = money(findtext(inf_vals, "vISSQN"))
        v_total_ret = money(findtext(inf_vals, "vTotalRet"))
        v_liq      = money(findtext(inf_vals, "vLiq"))

        # â”€â”€ ISS retention â”€â”€
        tp_ret_iss = findtext(trib_mun, "tpRetISSQN")  # 1=nÃ£o retido, 2=retido
        v_iss_ret  = money(findtext(trib_mun, "vISSQNRet"))
        # fallback: if tpRetISSQN==2 and vISSQNRet not present, use vTotalRet
        if tp_ret_iss == "2" and v_iss_ret == 0.0:
            v_iss_ret = v_total_ret

        # -- Federal retentions --
        v_irrf  = money(findtext(ret_fed, "vRetIRRF"))
        v_csll  = money(findtext(ret_fed, "vRetCSLL"))
        v_inss  = money(findtext(ret_fed, "vRetINSS"))
        piscofins = trib_fed.find(tag("piscofins")) if trib_fed is not None else None
        tp_ret_pc = findtext(piscofins, "tpRetPisCofins") if piscofins is not None else None
        if tp_ret_pc == "1":
            v_pis    = money(findtext(piscofins, "vPis"))
            v_cofins = money(findtext(piscofins, "vCofins"))
        else:
            v_pis    = money(findtext(ret_fed, "vRetPIS"))
            v_cofins = money(findtext(ret_fed, "vRetCOFINS"))

        # chave de acesso from filename (most reliable)
        chave = path.stem.replace("NFSe_", "", 1)
        # first 14 digits of chave after CNPJ prefix = not needed; use full stem

        return {
            "chave":       path.stem,       # full filename stem for PDF matching
            "xml_path":    path,
            "toma_cnpj":   toma_cnpj,
            "toma_nome":   toma_nome,
            "prest_cnpj":  prest_cnpj,
            "prest_nome":  prest_nome,
            "n_nfse":      n_nfse,
            "d_compet":    d_compet,
            "dh_proc":     dh_proc,
            "xloc":        xloc,
            "xdesc":       xdesc,
            "v_serv":      v_serv,
            "v_bc":        v_bc,
            "v_issqn":     v_issqn,
            "tp_ret_iss":  tp_ret_iss,
            "v_iss_ret":   v_iss_ret,
            "v_irrf":      v_irrf,
            "v_csll":      v_csll,
            "v_pis":       v_pis,
            "v_cofins":    v_cofins,
            "v_inss":      v_inss,
            "v_total_ret": v_total_ret,
            "v_liq":       v_liq,
        }
    except Exception as e:
        print(f"  [WARN] Could not parse {path.name}: {e}")
        return None


# â”€â”€ Build PDF index â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def build_pdf_index(pdf_dirs: list[Path]) -> dict[str, Path]:
    """stem (without extension) â†’ Path"""
    index = {}
    for d in pdf_dirs:
        for pdf in d.rglob("*.pdf"):
            index[pdf.stem] = pdf
    return index


# â”€â”€ Sanitize folder name â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def safe_name(name: str) -> str:
    for ch in r'\/:*?"<>|':
        name = name.replace(ch, "_")
    return name.strip().rstrip('. ')


# â”€â”€ Excel report â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
SUBHDR_FILL  = PatternFill("solid", start_color="2E75B6")
RET_FILL     = PatternFill("solid", start_color="FFF2CC")
ALT_FILL     = PatternFill("solid", start_color="DEEAF1")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")
TOTAL_FILL   = PatternFill("solid", start_color="D6E4F0")

WHITE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
REG_FONT     = Font(name="Arial", size=10)
MONEY_FMT    = '#,##0.00'
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def write_report(notas: list[dict], out_dir: Path):
    wb = Workbook()

    # â”€â”€ Sheet 1: Consolidado â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws_all = wb.active
    ws_all.title = "Consolidado"

    cols = [
        ("CNPJ Tomador",    16), ("Tomador",          40),
        ("CNPJ Prestador",  16), ("Prestador",         40),
        ("NÂº NFS-e",        10), ("CompetÃªncia",       14),
        ("MunicÃ­pio",        18), ("DescriÃ§Ã£o ServiÃ§o", 50),
        ("Vr. ServiÃ§o",     14), ("Base CÃ¡lculo",      14),
        ("ISSQN",           12), ("ISS Retido?",       12),
        ("Vr. ISS Ret.",    14), ("IRRF",              12),
        ("CSLL",            12), ("PIS",               12),
        ("COFINS",          12), ("INSS",              12),
        ("Total RetenÃ§Ãµes", 16), ("Vr. LÃ­quido",       14),
    ]

    header = [c[0] for c in cols]
    ws_all.append(header)
    for i, (_, width) in enumerate(cols, 1):
        cell = ws_all.cell(row=1, column=i)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = thin_border()
        ws_all.column_dimensions[cell.column_letter].width = width

    ws_all.row_dimensions[1].height = 22
    ws_all.freeze_panes = "A2"

    fields = [
        "toma_cnpj", "toma_nome", "prest_cnpj", "prest_nome",
        "n_nfse", "d_compet", "xloc", "xdesc",
        "v_serv", "v_bc", "v_issqn",
    ]
    money_cols = {9, 10, 11, 13, 14, 15, 16, 17, 18, 19, 20}  # 1-indexed

    for row_i, n in enumerate(notas, 2):
        fill = ALT_FILL if row_i % 2 == 0 else WHITE_FILL
        iss_ret_label = "Sim" if n["tp_ret_iss"] == "2" else "NÃ£o"
        row = [
            n["toma_cnpj"], n["toma_nome"],
            n["prest_cnpj"], n["prest_nome"],
            n["n_nfse"], n["d_compet"], n["xloc"], n["xdesc"],
            n["v_serv"], n["v_bc"], n["v_issqn"],
            iss_ret_label,
            n["v_iss_ret"], n["v_irrf"], n["v_csll"],
            n["v_pis"], n["v_cofins"], n["v_inss"],
            n["v_total_ret"], n["v_liq"],
        ]
        ws_all.append(row)
        for col_i, val in enumerate(row, 1):
            cell = ws_all.cell(row=row_i, column=col_i)
            cell.font = REG_FONT
            cell.border = thin_border()
            cell.alignment = CENTER if col_i in {1, 3, 5, 6, 12} else LEFT
            if col_i in money_cols:
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")
            # highlight retention row if any retention exists
            if n["v_total_ret"] > 0:
                cell.fill = RET_FILL
            else:
                cell.fill = fill

    # totals row
    last = ws_all.max_row
    ws_all.append([])
    tr = last + 1
    ws_all.cell(tr, 1, "TOTAIS").font = BOLD_FONT
    ws_all.cell(tr, 1).fill = TOTAL_FILL
    for col_i in money_cols:
        col_letter = ws_all.cell(1, col_i).column_letter
        c = ws_all.cell(tr, col_i)
        c.value = f"=SUM({col_letter}2:{col_letter}{last})"
        c.font = BOLD_FONT
        c.number_format = MONEY_FMT
        c.fill = TOTAL_FILL
        c.alignment = Alignment(horizontal="right", vertical="center")

    # â”€â”€ Sheet 2: SÃ³ RetenÃ§Ãµes â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws_ret = wb.create_sheet("RetenÃ§Ãµes")
    ws_ret.append(header)
    for i, (_, width) in enumerate(cols, 1):
        cell = ws_ret.cell(row=1, column=i)
        cell.font = WHITE_FONT
        cell.fill = SUBHDR_FILL
        cell.alignment = CENTER
        cell.border = thin_border()
        ws_ret.column_dimensions[cell.column_letter].width = width
    ws_ret.row_dimensions[1].height = 22
    ws_ret.freeze_panes = "A2"

    ret_notas = [n for n in notas if n["v_total_ret"] > 0]
    for row_i, n in enumerate(ret_notas, 2):
        iss_ret_label = "Sim" if n["tp_ret_iss"] == "2" else "NÃ£o"
        row = [
            n["toma_cnpj"], n["toma_nome"],
            n["prest_cnpj"], n["prest_nome"],
            n["n_nfse"], n["d_compet"], n["xloc"], n["xdesc"],
            n["v_serv"], n["v_bc"], n["v_issqn"],
            iss_ret_label,
            n["v_iss_ret"], n["v_irrf"], n["v_csll"],
            n["v_pis"], n["v_cofins"], n["v_inss"],
            n["v_total_ret"], n["v_liq"],
        ]
        ws_ret.append(row)
        for col_i, val in enumerate(row, 1):
            cell = ws_ret.cell(row=row_i, column=col_i)
            cell.font = REG_FONT
            cell.border = thin_border()
            cell.fill = RET_FILL if row_i % 2 == 0 else WHITE_FILL
            cell.alignment = CENTER if col_i in {1, 3, 5, 6, 12} else LEFT
            if col_i in money_cols:
                cell.number_format = MONEY_FMT
                cell.alignment = Alignment(horizontal="right", vertical="center")

    if ret_notas:
        last = ws_ret.max_row
        ws_ret.append([])
        tr = last + 1
        ws_ret.cell(tr, 1, "TOTAIS").font = BOLD_FONT
        ws_ret.cell(tr, 1).fill = TOTAL_FILL
        for col_i in money_cols:
            col_letter = ws_ret.cell(1, col_i).column_letter
            c = ws_ret.cell(tr, col_i)
            c.value = f"=SUM({col_letter}2:{col_letter}{last})"
            c.font = BOLD_FONT
            c.number_format = MONEY_FMT
            c.fill = TOTAL_FILL
            c.alignment = Alignment(horizontal="right", vertical="center")

    # â”€â”€ Sheet 3: Por Empresa â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws_emp = wb.create_sheet("Por Empresa")
    ws_emp.column_dimensions["A"].width = 40
    ws_emp.column_dimensions["B"].width = 16
    ws_emp.column_dimensions["C"].width = 10
    ws_emp.column_dimensions["D"].width = 16
    ws_emp.column_dimensions["E"].width = 16
    ws_emp.column_dimensions["F"].width = 16
    ws_emp.column_dimensions["G"].width = 16
    ws_emp.column_dimensions["H"].width = 16
    ws_emp.column_dimensions["I"].width = 16
    ws_emp.column_dimensions["J"].width = 16

    hdr = ["Empresa (Tomador)", "CNPJ", "Qtd Notas",
           "Total ServiÃ§os", "Total ISS", "ISS Retido",
           "IRRF", "CSLL", "PIS", "COFINS", "INSS", "Total RetenÃ§Ãµes"]
    ws_emp.append(hdr)
    for i in range(1, len(hdr) + 1):
        cell = ws_emp.cell(1, i)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = thin_border()

    # group by tomador
    by_company = defaultdict(list)
    for n in notas:
        by_company[(n["toma_cnpj"], n["toma_nome"])].append(n)

    for row_i, ((cnpj, nome), ns_list) in enumerate(
        sorted(by_company.items(), key=lambda x: x[0][1]), 2
    ):
        qtd       = len(ns_list)
        t_serv    = sum(n["v_serv"] for n in ns_list)
        t_iss     = sum(n["v_issqn"] for n in ns_list)
        t_iss_ret = sum(n["v_iss_ret"] for n in ns_list)
        t_irrf    = sum(n["v_irrf"] for n in ns_list)
        t_csll    = sum(n["v_csll"] for n in ns_list)
        t_pis     = sum(n["v_pis"] for n in ns_list)
        t_cofins  = sum(n["v_cofins"] for n in ns_list)
        t_inss    = sum(n["v_inss"] for n in ns_list)
        t_ret     = sum(n["v_total_ret"] for n in ns_list)

        row = [nome, cnpj, qtd, t_serv, t_iss, t_iss_ret,
               t_irrf, t_csll, t_pis, t_cofins, t_inss, t_ret]
        ws_emp.append(row)
        fill = ALT_FILL if row_i % 2 == 0 else WHITE_FILL
        for col_i, _ in enumerate(row, 1):
            cell = ws_emp.cell(row_i, col_i)
            cell.font = REG_FONT
            cell.border = thin_border()
            cell.fill = fill
            cell.alignment = LEFT if col_i <= 2 else Alignment(horizontal="right", vertical="center")
            if col_i >= 4:
                cell.number_format = MONEY_FMT

    out_path = out_dir / "NFSe_Retencoes_Report.xlsx"
    wb.save(out_path)
    return out_path


# â”€â”€ Main â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def parse_nota_xml(path):
    """Parse Questor flat <Nota> format."""
    try:
        r = ET.parse(path).getroot()
        if r.tag != 'Nota':
            return None
        def ft(t): return (r.findtext(t) or '').strip()
        def money(s):
            try: return float((s or '0').replace(',', '.'))
            except: return 0.0
        if ft('CANCELADO').upper() == 'TRUE':
            return None
        toma_cnpj = ft('CPFCNPJ_TOMADOR')
        toma_nome = ft('NOME_TOMADOR')
        prest_cnpj = ft('CPFCNPJ_PRESTADOR')
        prest_nome = ft('NOME_PRESTADOR')
        chave = ft('CHAVE') or path.stem
        v_irrf  = money(ft('VL_IR'))
        v_inss  = money(ft('VL_INSS'))
        v_iss_ret = money(ft('ISS_RETIDO'))
        # PIS/COFINS/CSLL only retained when VL_IR is also retained
        if v_irrf > 0:
            v_csll   = money(ft('VL_CSLL'))
            v_pis    = money(ft('VL_PIS'))
            v_cofins = money(ft('VL_COFINS'))
        else:
            v_csll = v_pis = v_cofins = 0.0
        return {
            'chave':      f'NFSe_{toma_cnpj}_{chave}',
            'xml_path':   path,
            'toma_cnpj':  toma_cnpj,
            'toma_nome':  toma_nome,
            'prest_cnpj': prest_cnpj,
            'prest_nome': prest_nome,
            'n_nfse':     ft('N_DA_NFSE'),
            'd_compet':   ft('COMPETENCIA'),
            'dh_proc':    ft('DATA_EMISSAO'),
            'xloc':       ft('NOME_CIDADE_PRESTADOR'),
            'xdesc':      ft('CODIGO_SERVICO'),
            'v_serv':     money(ft('VALOR_LIQUIDO')),
            'v_bc':       money(ft('VALOR_LIQUIDO')),
            'v_issqn':    money(ft('VL_ISS')),
            'tp_ret_iss': '1' if v_iss_ret > 0 else '2',
            'v_iss_ret':  v_iss_ret,
            'v_irrf':     v_irrf,
            'v_csll':     v_csll,
            'v_pis':      v_pis,
            'v_cofins':   v_cofins,
            'v_inss':     v_inss,
            'v_total_ret': v_irrf + v_csll + v_pis + v_cofins + v_inss + v_iss_ret,
            'v_liq':      money(ft('VALOR_LIQUIDO')),
        }
    except Exception as e:
        print(f'  [WARN] Could not parse Nota {path.name}: {e}')
        return None


def parse_nota_xml(path):
    """Parse Questor flat <Nota> format."""
    try:
        r = ET.parse(path).getroot()
        if r.tag != 'Nota':
            return None
        def ft(t): return (r.findtext(t) or '').strip()
        def money(s):
            try: return float((s or '0').replace(',', '.'))
            except: return 0.0
        if ft('CANCELADO').upper() == 'TRUE':
            return None
        toma_cnpj = ft('CPFCNPJ_TOMADOR')
        toma_nome = ft('NOME_TOMADOR')
        prest_cnpj = ft('CPFCNPJ_PRESTADOR')
        prest_nome = ft('NOME_PRESTADOR')
        chave = ft('CHAVE') or path.stem
        v_irrf  = money(ft('VL_IR'))
        v_inss  = money(ft('VL_INSS'))
        v_iss_ret = money(ft('ISS_RETIDO'))
        # PIS/COFINS/CSLL only retained when VL_IR is also retained
        if v_irrf > 0:
            v_csll   = money(ft('VL_CSLL'))
            v_pis    = money(ft('VL_PIS'))
            v_cofins = money(ft('VL_COFINS'))
        else:
            v_csll = v_pis = v_cofins = 0.0
        return {
            'chave':      f'NFSe_{toma_cnpj}_{chave}',
            'xml_path':   path,
            'toma_cnpj':  toma_cnpj,
            'toma_nome':  toma_nome,
            'prest_cnpj': prest_cnpj,
            'prest_nome': prest_nome,
            'n_nfse':     ft('N_DA_NFSE'),
            'd_compet':   ft('COMPETENCIA'),
            'dh_proc':    ft('DATA_EMISSAO'),
            'xloc':       ft('NOME_CIDADE_PRESTADOR'),
            'xdesc':      ft('CODIGO_SERVICO'),
            'v_serv':     money(ft('VALOR_LIQUIDO')),
            'v_bc':       money(ft('VALOR_LIQUIDO')),
            'v_issqn':    money(ft('VL_ISS')),
            'tp_ret_iss': '1' if v_iss_ret > 0 else '2',
            'v_iss_ret':  v_iss_ret,
            'v_irrf':     v_irrf,
            'v_csll':     v_csll,
            'v_pis':      v_pis,
            'v_cofins':   v_cofins,
            'v_inss':     v_inss,
            'v_total_ret': v_irrf + v_csll + v_pis + v_cofins + v_inss + v_iss_ret,
            'v_liq':      money(ft('VALOR_LIQUIDO')),
        }
    except Exception as e:
        print(f'  [WARN] Could not parse Nota {path.name}: {e}')
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xml-dirs", nargs="+", required=True,
                    help="One or more directories containing NFSe XML files (searched recursively)")
    ap.add_argument("--pdf-dirs", nargs="+", default=[],
                    help="One or more directories containing DANFSe PDF files")
    ap.add_argument("--out-dir",  required=True,
                    help="Output directory (will be created if needed)")
    ap.add_argument("--only-retencoes", action="store_true",
                    help="Only copy PDFs / create folders for notas with actual retencoes")
    args = ap.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # â”€â”€ collect XMLs â”€â”€
    xml_paths = []
    for d in args.xml_dirs:
        xml_paths.extend(Path(d).rglob("*.xml"))
    print(f"Found {len(xml_paths)} XML files.")

    # â”€â”€ parse â”€â”€
    notas = []
    for p in sorted(xml_paths):
        n = parse_xml(p)
        if n is None:
            n = parse_nota_xml(p)
        if n:
            notas.append(n)
    print(f"Parsed {len(notas)} valid notas.")

    if not notas:
        print("No notas parsed â€” check your --xml-dirs paths.")
        sys.exit(1)

    # â”€â”€ build PDF index â”€â”€
    pdf_index = build_pdf_index([Path(d) for d in args.pdf_dirs])
    print(f"Indexed {len(pdf_index)} PDFs.")

    # â”€â”€ create per-company Retencao folders â”€â”€
    companies_dir = out_dir / "Empresas"
    companies_dir.mkdir(exist_ok=True)

    # decide which notas get folders
    notas_for_folders = (
        [n for n in notas if n["v_irrf"] > 0 or n["v_csll"] > 0 or n["v_inss"] > 0 or n["v_pis"] > 0 or n["v_cofins"] > 0]
        if args.only_retencoes else notas
    )

    copied_xml = 0
    copied_pdf = 0
    missing_pdf = []

    by_company = defaultdict(list)
    for n in notas_for_folders:
        key = ''.join(filter(str.isdigit, n['toma_cnpj']))
        by_company[key].append(n)

    all_by_company = defaultdict(list)
    for n in notas:
        key = ''.join(filter(str.isdigit, n['toma_cnpj']))
        all_by_company[key].append(n)

    ret_by_company = defaultdict(list)
    for n in notas_for_folders:
        key = ''.join(filter(str.isdigit, n['toma_cnpj']))
        ret_by_company[key].append(n)

    for company_key, all_ns in all_by_company.items():
        from collections import Counter
        best_name = Counter(safe_name(n['toma_nome']) for n in all_ns).most_common(1)[0][0]
        folder_name = f"{best_name} {company_key}"
        base = companies_dir / folder_name

        todas_xml = base / "Todas as Notas" / "xml"
        todas_pdf = base / "Todas as Notas" / "pdf"
        for d in [todas_xml, todas_pdf]:
            d.mkdir(parents=True, exist_ok=True)

        for n in all_ns:
            dest = todas_xml / n["xml_path"].name
            if not dest.exists():
                shutil.copy2(n["xml_path"], dest)
                copied_xml += 1
            stem = n["chave"]
            if stem in pdf_index:
                dest = todas_pdf / pdf_index[stem].name
                if not dest.exists():
                    shutil.copy2(pdf_index[stem], dest)
                    copied_pdf += 1
            else:
                missing_pdf.append(stem)

        ns_ret_list = ret_by_company.get(company_key, [])
        if ns_ret_list:
            ret_xml = base / "Retencao Federal" / "xml"
            ret_pdf = base / "Retencao Federal" / "pdf"
            for d in [ret_xml, ret_pdf]:
                d.mkdir(parents=True, exist_ok=True)
            for n in ns_ret_list:
                dest = ret_xml / n["xml_path"].name
                if not dest.exists():
                    shutil.copy2(n["xml_path"], dest)
                stem = n["chave"]
                if stem in pdf_index:
                    dest = ret_pdf / pdf_index[stem].name
                    if not dest.exists():
                        shutil.copy2(pdf_index[stem], dest)

    print(f"Copied {copied_xml} XMLs and {copied_pdf} PDFs to company Retencao folders.")
    if missing_pdf:
        print(f"  [{len(missing_pdf)} PDFs not found in --pdf-dirs]")
        (out_dir / "missing_pdfs.txt").write_text("\n".join(missing_pdf))
        print(f"  â†’ List saved to: {out_dir / 'missing_pdfs.txt'}")

    # â”€â”€ write Excel report â”€â”€
    print("Generating Excel report...")
    report_path = write_report(notas, out_dir)
    print(f"\nDone. Report: {report_path}")
    print(f"Company folders: {companies_dir}")


if __name__ == "__main__":
    main()
